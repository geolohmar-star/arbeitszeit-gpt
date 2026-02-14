# schichtplan/views.py - ANGEPASST für MA1-MA15 mit Präferenzen
"""
Views für das Schichtplan-Modul
ANGEPASST: Filtert nur Mitarbeiter mit Kennung MA1-MA15

"""
from django.db.models import Count, Q
from django.db.models.functions import Length  # ← Für Sortierung
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count
from django.http import JsonResponse, HttpResponse
from django.views.generic.edit import CreateView
from django.urls import reverse_lazy
from django.db import transaction
from django import forms
from collections import defaultdict
from django.http import StreamingHttpResponse
import time
from django.utils import timezone
from datetime import datetime, timedelta, date
from calendar import monthrange



from datetime import timedelta
from calendar import day_name
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from docx import Document
from weasyprint import HTML

# Models
from arbeitszeit.models import Mitarbeiter, MonatlicheArbeitszeitSoll
from .models import Schichtplan, Schicht, Schichttyp, SchichtwunschPeriode, Schichtwunsch, SchichtplanAenderung, SchichtplanSnapshot, SchichtplanSnapshotSchicht


# Forms
from .forms import ExcelImportForm, SchichtplanForm, SchichtForm

# Services
from .services import SchichtplanGenerator

# Utils

try:
    from .utils.excel_import import SchichtplanImporter
except ImportError:
    try:
        from .utils.xls_importer import SchichtplanImporter
    except ImportError:
        SchichtplanImporter = None

#Wunschplan
from datetime import timedelta, timedelta, date
from calendar import monthrange



def _ist_stunden_urlaub_krank(ma, daten_urlaub_krank, feiertage_set=None):
    """
    Addiert zu den Ist-Stunden: Für jeden Urlaub-/Krank-/gar_nichts-Tag (nur Mo–Fr, OHNE Feiertage)
    die Tagesstunden aus der gültigen Arbeitszeitvereinbarung (Wochenstunden / 5).
    
    NEU: Berücksichtigt NRW-Feiertage - diese werden NICHT als Arbeitstage gezählt!
    """
    if not daten_urlaub_krank:
        return 0.0
    
    stunden = 0.0
    
    # Falls keine Feiertage übergeben, berechne sie
    if feiertage_set is None:
        min_datum = min(daten_urlaub_krank)
        max_datum = max(daten_urlaub_krank)
        feiertage_set, _ = get_configured_feiertage(min_datum, max_datum, region='nrw')
    
    for d in daten_urlaub_krank:
        # Wochenende überspringen
        if d.weekday() >= 5:  # Sa, So = kein Arbeitstag
            continue
        
        # NEU: Feiertage überspringen
        if d in feiertage_set:
            continue
        
        # Tagesstunden aus Vereinbarung holen
        v = ma.get_aktuelle_vereinbarung(d)
        if v and v.wochenstunden is not None:
            stunden += float(v.wochenstunden) / 5.0
    
    return stunden

try:
    from dateutil.easter import easter
except ImportError:
    easter = None 

 

# ============================================================================
# HELPER FUNKTIONEN
# ============================================================================

def generate_schichtplan(request):
    if request.method == 'POST':
        def event_stream():
            yield "data: Starting solver...\n\n"
            # Dein Solver hier
            result = solver.solve()
            yield f"data: Complete\n\n"
        
        return StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    if status == cp_model.OPTIMAL:
        messages.success(request, "✅ Optimaler Schichtplan erstellt!")
    elif status == cp_model.FEASIBLE:
        messages.warning(request, "⚠️ Gültiger Schichtplan erstellt (nicht optimal, aber verwendbar)")
    else:
        messages.error(request, "❌ Kein gültiger Schichtplan möglich")
    
def ist_schichtplaner(user):
    # DEBUG: Zeigt im Terminal an, wer gerade prüft
    if user.is_anonymous:
        return False
    
    # 1. Admin/Superuser (funktioniert immer)
    if user.is_superuser or user.is_staff:
        return True
    
    # 2. Gruppen-Prüfung (WICHTIG: Name muss exakt stimmen)
    # Wir prüfen, ob der User in der Gruppe "Schichtplaner" ist
    ergebnis = user.groups.filter(name='Schichtplaner').exists()
    
    # DEBUG-Ausgabe für dich im Terminal:
    print(f"--- Berechtigungs-Check für {user.username} ---")
    print(f"Ist Staff: {user.is_staff}")
    print(f"In Gruppe Schichtplaner: {ergebnis}")
    
    return ergebnis


def ist_kongos_mitarbeiter(user):
    """True wenn User ein Mitarbeiter der Abteilung Kongos ist (Zugriff auf veröffentlichte Pläne)."""
    if user.is_anonymous:
        return False
    if not hasattr(user, 'mitarbeiter'):
        return False
    try:
        return (user.mitarbeiter.abteilung or '').strip().lower() == 'kongos'
    except Exception:
        return False


def darf_schichtplan_sehen(user, schichtplan):
    """True wenn User den Plan sehen darf: Schichtplaner immer, Kongos nur bei veröffentlichtem Plan."""
    if ist_schichtplaner(user):
        return True
    if schichtplan.status == 'veroeffentlicht' and ist_kongos_mitarbeiter(user):
        return True
    return False


def darf_plan_bearbeiten(user, schichtplan):
    """True wenn User den Plan bearbeiten darf (Schichten tauschen/löschen/hinzufügen): Schichtplaner immer, Kongos bei veröffentlichtem Plan (alle Änderungen werden protokolliert)."""
    if ist_schichtplaner(user):
        return True
    if schichtplan.status == 'veroeffentlicht' and ist_kongos_mitarbeiter(user):
        return True
    return False


#def ist_schichtplaner(user):
#    """Prüft, ob der User Schichtplaner-Rechte hat"""
#    if user.is_staff or user.is_superuser:
#        return True
#    
#    if hasattr(user, 'mitarbeiter'):
#        return user.mitarbeiter.rolle == 'schichtplaner'
#    
#    return False


def get_planbare_mitarbeiter():
    """
    Gibt nur Mitarbeiter zurück, die für Schichtplanung relevant sind.
    
    Kriterien:
    - Aktiv = True
    - Kennung = MA1 bis MA15
    - Verfügbarkeit != 'dauerkrank'
    """
    gueltige_kennungen = [f'MA{i}' for i in range(1, 16)]  # MA1 bis MA15
    
    return Mitarbeiter.objects.filter(
        aktiv=True,
        schichtplan_kennung__in=gueltige_kennungen
    ).exclude(
        verfuegbarkeit='dauerkrank'
    ).select_related('user')


def get_nrw_feiertage(start_date, end_date):
    """
    Feiertage in NRW (Nordrhein-Westfalen) im angegebenen Zeitraum.
    Returns: (set of date, dict date -> Bezeichnung)
    """
    feiertage_set = set()
    feiertage_namen = {}
    if not easter:
        return feiertage_set, feiertage_namen

    start = start_date if isinstance(start_date, date) else start_date
    end = end_date if isinstance(end_date, date) else end_date
    years = {start.year, end.year}
    if start.month == 12 and end.month == 1:
        years.add(start.year + 1)

    for year in years:
        # Fixe Feiertage
        for d, name in [
            (date(year, 1, 1), "Neujahr"),
            (date(year, 5, 1), "Tag der Arbeit"),
            (date(year, 10, 3), "Tag der Deutschen Einheit"),
            (date(year, 11, 1), "Allerheiligen"),
            (date(year, 12, 25), "1. Weihnachtsfeiertag"),
            (date(year, 12, 26), "2. Weihnachtsfeiertag"),
        ]:
            if start <= d <= end:
                feiertage_set.add(d)
                feiertage_namen[d] = name

        # Bewegliche Feiertage (Ostern)
        ostersonntag = easter(year)
        for delta, name in [
            (-2, "Karfreitag"),
            (1, "Ostermontag"),
            (39, "Christi Himmelfahrt"),
            (50, "Pfingstmontag"),
            (60, "Fronleichnam"),
        ]:
            d = ostersonntag + timedelta(days=delta)
            if start <= d <= end:
                feiertage_set.add(d)
                feiertage_namen[d] = name

    return feiertage_set, feiertage_namen


def get_configured_feiertage(start_date, end_date, region='all'):
    """
    Lädt Feiertage aus der Datenbank (RegionalerFeiertag-Modell).
    Falls keine Feiertage konfiguriert: Fallback auf get_nrw_feiertage()

    Args:
        start_date, end_date: Zeitraum
        region: 'nrw', 'bayern', 'bw', 'all', etc.

    Returns: (set of dates, dict date → name)
    """
    from schichtplan.models import RegionalerFeiertag
    from django.db.models import Q

    feiertage_set = set()
    feiertage_namen = {}

    if not easter:
        return feiertage_set, feiertage_namen

    # 1. Hole alle aktiven Feiertage für die Region
    feiertage = RegionalerFeiertag.objects.filter(aktiv=True).filter(
        Q(region='all') | Q(region=region)
    )

    # Falls keine Feiertage konfiguriert: Fallback auf get_nrw_feiertage()
    if not feiertage.exists():
        return get_nrw_feiertage(start_date, end_date)

    # 2. Verarbeite feste Feiertage
    for f in feiertage.filter(typ='fest'):
        if f.monat and f.tag:
            for year in {start_date.year, end_date.year}:
                try:
                    d = date(year, f.monat, f.tag)
                    if start_date <= d <= end_date:
                        feiertage_set.add(d)
                        feiertage_namen[d] = f.name
                except ValueError:  # Invalid date (z.B. 31. Februar)
                    pass

    # 3. Verarbeite Ostern-relative Feiertage
    for f in feiertage.filter(typ='ostern_relativ'):
        if f.ostern_offset is not None:
            for year in {start_date.year, end_date.year}:
                ostersonntag = easter(year)
                d = ostersonntag + timedelta(days=f.ostern_offset)
                if start_date <= d <= end_date:
                    feiertage_set.add(d)
                    feiertage_namen[d] = f.name

    return feiertage_set, feiertage_namen


def _get_monat_pulldown_choices(anzahl_monate=24):
    """Erstellt Choices für Monats-Pulldown: (iso_date, 'Januar 2026'), ... ab aktuellem Monat."""
    MONATE_DE = ('', 'Januar', 'Februar', 'März', 'April', 'Mai', 'Juni',
                 'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember')
    heute = date.today()
    d = date(heute.year, heute.month, 1)
    choices = [('', '–– Monat wählen ––')]
    for _ in range(anzahl_monate):
        label = f"{MONATE_DE[d.month]} {d.year}"
        choices.append((d.isoformat(), label))
        if d.month == 12:
            d = date(d.year + 1, 1, 1)
        else:
            d = date(d.year, d.month + 1, 1)
    return choices


class WunschPeriodeCreateView(CreateView):
    model = SchichtwunschPeriode
    template_name = 'schichtplan/periode_form.html'
    fields = ['name', 'fuer_monat', 'eingabe_start', 'eingabe_ende', 'status']
    success_url = reverse_lazy('schichtplan:wunsch_perioden_liste')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        
        # 1. Für Monat: Pulldown (Monat/Jahr)
        form.fields['fuer_monat'].widget = forms.Select(
            choices=_get_monat_pulldown_choices(),
            attrs={'class': 'form-select'}
        )
        form.fields['fuer_monat'].help_text = 'Wählen Sie den Monat, für den die Wünsche gelten.'
        
        # 2. Datum mit Uhrzeit (Eingabezeitraum)
        form.fields['eingabe_start'].widget = forms.DateTimeInput(
            attrs={'type': 'datetime-local', 'class': 'form-control'}
        )
        form.fields['eingabe_ende'].widget = forms.DateTimeInput(
            attrs={'type': 'datetime-local', 'class': 'form-control'}
        )
        
        form.fields['name'].widget.attrs.update({'class': 'form-control'})
        form.fields['status'].widget.attrs.update({'class': 'form-select'})
        
        return form

    def form_valid(self, form):
        form.instance.erstellt_von = self.request.user
        return super().form_valid(form)


# ============================================================================
# CLASS-BASED VIEW: Schichtplan erstellen
# ============================================================================

class SchichtplanCreateView(CreateView):
    """
    Erstellt einen neuen Schichtplan.
    Verwendet nur Mitarbeiter mit Kennung MA1-MA15.
    """
    
    model = Schichtplan
    form_class = SchichtplanForm
    template_name = 'schichtplan/schichtplan_erstellen.html'
    success_url = reverse_lazy('schichtplan:dashboard')

    def dispatch(self, request, *args, **kwargs):
        """Berechtigungsprüfung"""
        if not ist_schichtplaner(request.user):
            messages.error(request, "❌ Keine Berechtigung für diese Aktion.")
            return redirect('arbeitszeit:dashboard')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        """Speichern + Optional KI-Generierung"""
        
        ki_aktiviert = form.cleaned_data.get('vorschlag_generieren', False)
        
        try:
            with transaction.atomic():
                # 1. Schichtplan speichern
                self.object = form.save(commit=False)
                
                if hasattr(self.object, 'erstellt_von'):
                    self.object.erstellt_von = self.request.user
                
                self.object.save()
                
                print(f"[OK] Schichtplan '{self.object.name}' (ID={self.object.pk}) gespeichert")
                
                # 2. KI-Generierung (falls aktiviert)
                if ki_aktiviert:
                    self._generate_with_ai()
                else:
                    messages.success(
                        self.request,
                        f"✅ Plan '{self.object.name}' wurde leer angelegt."
                    )
            
            return redirect(self.get_success_url())
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            
            messages.error(
                self.request,
                f"❌ Fehler beim Erstellen des Plans: {str(e)}"
            )
            return self.form_invalid(form)

    def _generate_with_ai(self):
      
        print(f"[KI] Generierung fuer Plan '{self.object.name}' (ID={self.object.pk}) gestartet...")
        
        # 1. BASIS-CHECK: Mitarbeiter vorhanden?
        planbare_mitarbeiter = get_planbare_mitarbeiter()
        
        if not planbare_mitarbeiter.exists():
            messages.warning(
                self.request,
                "⚠️ Keine planbaren Mitarbeiter gefunden (MA1-MA15 aktiv/nicht dauerkrank)."
            )
            return
        
        # 2. BASIS-CHECK: Schichttypen vorhanden?
        required_types = ['T', 'N']
        existing_types = list(Schichttyp.objects.filter(kuerzel__in=required_types).values_list('kuerzel', flat=True))
        
        if len(existing_types) != len(required_types):
            missing = set(required_types) - set(existing_types)
            raise Exception(f"Schichttypen fehlen: {', '.join(missing)}. Bitte im Admin anlegen.")
        
        # 3. HINWEIS: Wünsche werden automatisch vom Generator geladen
        print(f"   [INFO] Wuensche werden automatisch aus DB geladen (falls vorhanden)")
        
        # 4. GENERATOR STARTEN (Genau ein Aufruf)
        try:
            # WICHTIG: Nur ein Positionsargument!
            # Alte Schichten löschen (falls Plan neu generiert wird)
            Schicht.objects.filter(schichtplan=self.object).delete()
            generator = SchichtplanGenerator(planbare_mitarbeiter)
            
            generator.generiere_vorschlag(self.object)
            
            schichten_anzahl = self.object.schichten.count()
            
            messages.success(
                self.request,
                f"Plan '{self.object.name}' wurde erfolgreich erstellt! "
                f"{schichten_anzahl} Schichten fuer {planbare_mitarbeiter.count()} "
                f"Mitarbeiter automatisch generiert."
            )
            print(f"[OK] Erfolg: {schichten_anzahl} Schichten generiert.")
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise Exception(f"Die KI-Generierung schlug fehl: {str(e)}")


# ============================================================================
# REST DER VIEWS (unverändert)
# ============================================================================

@login_required
def excel_import_view(request, pk):
    """Importiert Excel-Datei in einen bestehenden Schichtplan"""
    schichtplan = get_object_or_404(Schichtplan, pk=pk)

    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('schichtplan:dashboard')

    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)

        if form.is_valid():
            excel_file = request.FILES['excel_file']

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                for chunk in excel_file.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name

            try:
                if SchichtplanImporter is None:
                    raise Exception("Excel-Importer nicht verfügbar!")
                
                importer = SchichtplanImporter()
                importer.import_excel_mit_zuordnung(tmp_path, schichtplan)

                messages.success(request, "✅ Excel-Datei erfolgreich importiert!")
                return redirect('schichtplan:detail', pk=schichtplan.pk)

            except Exception as e:
                messages.error(request, f"❌ Fehler beim Import: {str(e)}")

    else:
        form = ExcelImportForm()

    return render(request, 'schichtplan/excel_import.html', {
        'form': form,
        'schichtplan': schichtplan
    })


@login_required
def excel_analyse_view(request):
    """Analysiert Excel-Datei vor dem Import"""
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('arbeitszeit:dashboard')
    
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        
        if form.is_valid():
            stats = {
                'info': 'Analyse-Funktion noch nicht implementiert'
            }
            
            return render(request, 'schichtplan/excel_analyse.html', {
                'stats': stats,
                'form': form,
            })
    else:
        form = ExcelImportForm()
    
    return render(request, 'schichtplan/excel_analyse.html', {
        'form': form,
    })


@login_required
def planer_dashboard(request):
    """Dashboard für Schichtplaner. Kongos-Mitarbeiter sehen nur veröffentlichte Pläne (Lesezugriff)."""
    is_planer = ist_schichtplaner(request.user)
    is_kongos = ist_kongos_mitarbeiter(request.user)

    if not is_planer and not is_kongos:
        messages.error(request, "❌ Zugriff verweigert. Dieser Bereich ist nur für Schichtplaner oder Abteilung Kongos.")
        return redirect('schichtplan:wunsch_kalender_aktuell')

    if is_planer:
        schichtplaene = Schichtplan.objects.all().order_by('-start_datum')
        veroeffentlichte_plaene = schichtplaene.filter(status='veroeffentlicht')
        planbare_ma = get_planbare_mitarbeiter()
        context = {
            'is_planer': True,
            'is_kongos': False,
            'aktive_plaene': schichtplaene.filter(status='veroeffentlicht').count(),
            'entwuerfe': schichtplaene.filter(status='entwurf').count(),
            'zur_genehmigung_count': schichtplaene.filter(status='zur_genehmigung').count(),
            'schichtplaene': schichtplaene,
            'veroeffentlichte_plaene': veroeffentlichte_plaene,
            'mitarbeiter_gesamt': planbare_ma.count(),
            'mitarbeiter_zugeordnet': planbare_ma.exclude(schichtplan_kennung='').count(),
        }
    else:
        # Kongos: nur veröffentlichte Pläne anzeigen
        schichtplaene = Schichtplan.objects.filter(status='veroeffentlicht').order_by('-start_datum')
        context = {
            'is_planer': False,
            'is_kongos': True,
            'schichtplaene': schichtplaene,
            'aktive_plaene': schichtplaene.count(),
            'entwuerfe': 0,
            'zur_genehmigung_count': 0,
        }

    return render(request, 'schichtplan/planer_dashboard.html', context)


@login_required
def mitarbeiter_uebersicht(request):
    """Mitarbeiter-Übersicht für Schichtplaner"""
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('arbeitszeit:dashboard')
    
    # Zeige nur planbare Mitarbeiter
    mitarbeiter = get_planbare_mitarbeiter()
    
    stats = {
        'gesamt': mitarbeiter.count(),
        'zugeordnet': mitarbeiter.exclude(schichtplan_kennung='').count(),
        'nicht_zugeordnet': mitarbeiter.filter(schichtplan_kennung='').count(),
        'dauerkrank': Mitarbeiter.objects.filter(
            aktiv=True,
            verfuegbarkeit='dauerkrank'
        ).count(),
    }
    
    context = {
        'mitarbeiter': mitarbeiter,
        'mitarbeiter_liste': mitarbeiter,
        'stats': stats,
        'anzahl_siegburg': mitarbeiter.filter(standort='siegburg').count(),
        'anzahl_bonn': mitarbeiter.filter(standort='bonn').count(),
    }
    
    return render(request, 'arbeitszeit/mitarbeiter_uebersicht.html', context)


@login_required
def schichtplan_detail(request, pk):
    """Detail-Ansicht eines Schichtplans mit erweiterter Statistik. Kongos sehen nur veröffentlichte Pläne (lesend)."""
    schichtplan = get_object_or_404(Schichtplan, pk=pk)

    if not darf_schichtplan_sehen(request.user, schichtplan):
        messages.error(request, "❌ Keine Berechtigung für diesen Schichtplan.")
        return redirect('arbeitszeit:dashboard')

    # 1. Alle Schichten laden
    schichten = schichtplan.schichten.select_related(
        'mitarbeiter', 'schichttyp'
    ).order_by('datum', 'schichttyp__start_zeit')

    # 2. Kalender-Daten vorbereiten
    mitarbeiter_mapping = {ma.schichtplan_kennung: ma.vollname for ma in Mitarbeiter.objects.all()}
    kalender_daten = {}
    current_date = schichtplan.start_datum

    while current_date <= schichtplan.ende_datum:
        tag_schichten = [s for s in schichten if s.datum == current_date]
        kalender_daten[current_date] = {
            'datum': current_date,
            'wochentag': day_name[current_date.weekday()],
            'schichten': tag_schichten,
            'ist_wochenende': current_date.weekday() >= 5,
        }
        current_date += timedelta(days=1)

    # 3. STATISTIK BERECHNEN (KORRIGIERT)
    # ---------------------------------------------------------
    stats_list = []
    
    # WICHTIG: Wir holen ALLE relevanten Mitarbeiter, nicht nur die mit Schichten!
    alle_mitarbeiter = get_planbare_mitarbeiter() # Nutzt deine Helper-Funktion oben
    
    # Debug
    print(f"DEBUG: Berechne Statistik für {alle_mitarbeiter.count()} Mitarbeiter...")
    
    # Fallback-Stunden
    stunden_defaults = {'T': 12.25, 'N': 12.25, 'Z': 8.0}
    
    # NEU: Feiertage einmalig berechnen (für Urlaubs-Stunden-Berechnung)
    feiertage_set, _ = get_configured_feiertage(schichtplan.start_datum, schichtplan.ende_datum, region='nrw')

    # Urlaub/Krank/gar_nichts: Tage pro MA für Ist-Stunden-Zuschlag (Tagesstunden aus Vereinbarung)
    wuensche_uk = Schichtwunsch.objects.filter(
        mitarbeiter__in=alle_mitarbeiter,
        datum__gte=schichtplan.start_datum,
        datum__lte=schichtplan.ende_datum,
        wunsch__in=['urlaub', 'gar_nichts', 'krank']
    ).values_list('mitarbeiter_id', 'datum')
    urlaub_krank_pro_ma = defaultdict(list)
    for ma_id, datum in wuensche_uk:
        urlaub_krank_pro_ma[ma_id].append(datum)

    for ma in alle_mitarbeiter:
        # Schichten dieses Mitarbeiters aus der geladenen Liste filtern
        ma_schichten = [s for s in schichten if s.mitarbeiter_id == ma.id]
        
        c_t = 0
        c_n = 0
        c_z = 0
        ist_stunden = 0.0
        wochenenden_set = set()
        
        for s in ma_schichten:
            kuerzel = s.schichttyp.kuerzel
            
            if kuerzel == 'T': c_t += 1
            elif kuerzel == 'N': c_n += 1
            elif kuerzel == 'Z': c_z += 1
            
            # Stunden summieren
            # SPEZIALFALL MA7: Z-Schichten zählen wie Nachtschichten (12,25h)
            if ma.schichtplan_kennung == 'MA7' and kuerzel == 'Z':
                stunden = 12.25  # Gleich wie Nachtschicht
            # ALLE MIT ARBEITSZEITVEREINBARUNG: Z-Schichten aus Vereinbarung (Tagesarbeitszeit)
            elif kuerzel == 'Z':
                # Hole aktuelle Vereinbarung und berechne Tagesarbeitszeit
                vereinbarung = ma.get_aktuelle_vereinbarung(s.datum)
                if vereinbarung and vereinbarung.wochenstunden:
                    # Wochenstunden / 5 Tage = Tagesstunden
                    stunden = float(vereinbarung.wochenstunden) / 5.0
                else:
                    # Fallback: Schichttyp-Dauer wenn keine Vereinbarung
                    stunden = float(s.schichttyp.arbeitszeit_stunden) if s.schichttyp.arbeitszeit_stunden else stunden_defaults.get(kuerzel, 0)
            else:
                stunden = float(s.schichttyp.arbeitszeit_stunden) if s.schichttyp.arbeitszeit_stunden else stunden_defaults.get(kuerzel, 0)
            ist_stunden += stunden
            
            # Wochenende
            if s.datum.weekday() >= 5:
                iso_year, iso_week, _ = s.datum.isocalendar()
                wochenenden_set.add(f"{iso_year}-{iso_week}")

        # Urlaub/Krank/gar_nichts: Tagesstunden aus gültiger Vereinbarung zu Ist addieren (nur Mo–Fr, OHNE Feiertage)
        ist_stunden += _ist_stunden_urlaub_krank(ma, urlaub_krank_pro_ma.get(ma.id, []), feiertage_set)

        # --- SOLL-STUNDEN via Model-Methode ---
        try:
            # Nutzt deine Methode im Model
            soll_stunden = float(ma.get_soll_stunden_monat(
                schichtplan.start_datum.year, 
                schichtplan.start_datum.month
            ))
        except Exception as e:
            print(f"   [WARN] Fehler Soll-Stunden {ma.schichtplan_kennung}: {e}")
            soll_stunden = 160.0 # Fallback

        prozent = (ist_stunden / soll_stunden * 100) if soll_stunden > 0 else 0
        
        stats_list.append({
            'ma': ma,
            't': c_t,
            'n': c_n,
            'z': c_z,
            'we': len(wochenenden_set),
            'ist_stunden': ist_stunden,
            'soll_stunden': soll_stunden,
            'diff': ist_stunden - soll_stunden,
            'prozent': prozent
        })

    # Sortieren nach Kennung (MA1, MA2...)
    stats_list.sort(key=lambda x: (len(x['ma'].schichtplan_kennung or ''), x['ma'].schichtplan_kennung or ''))

    # Jahresübersicht: Pro Mitarbeiter die Summen T, N, Z, Wochenenden über alle Pläne des Jahres
    jahr = schichtplan.start_datum.year
    jahr_start = date(jahr, 1, 1)
    jahr_ende = date(jahr, 12, 31)
    plaene_jahr = Schichtplan.objects.filter(
        start_datum__lte=jahr_ende,
        ende_datum__gte=jahr_start
    )
    schichten_jahr = SchichtplanSnapshotSchicht.objects.filter(
        snapshot__schichtplan__in=plaene_jahr,
        datum__gte=jahr_start,
        datum__lte=jahr_ende
    ).select_related('schichttyp', 'mitarbeiter', 'snapshot__schichtplan')
    # Pro MA: t, n, z, we
    ma_jahr_stats = {ma.id: {'t': 0, 'n': 0, 'z': 0, 'we': 0} for ma in alle_mitarbeiter}
    for s in schichten_jahr:
        if s.mitarbeiter_id not in ma_jahr_stats:
            continue
        k = s.schichttyp.kuerzel
        if k == 'T':
            ma_jahr_stats[s.mitarbeiter_id]['t'] += 1
        elif k == 'N':
            ma_jahr_stats[s.mitarbeiter_id]['n'] += 1
        elif k == 'Z':
            ma_jahr_stats[s.mitarbeiter_id]['z'] += 1
        if s.datum.weekday() >= 5:
            ma_jahr_stats[s.mitarbeiter_id]['we'] += 1
    jahres_mitarbeiter = []
    for ma in alle_mitarbeiter:
        st = ma_jahr_stats.get(ma.id, {'t': 0, 'n': 0, 'z': 0, 'we': 0})
        jahres_mitarbeiter.append({
            'ma': ma,
            't': st['t'],
            'n': st['n'],
            'z': st['z'],
            'we': st['we'],
            'jahres_total': st['t'] + st['n'] + st['z'],
        })

    # Heatmap-Scoring je Spalte (T/N/WE), Teilzeit/Zusatz ausgeschlossen
    vollzeit_rows = [
        r for r in jahres_mitarbeiter
        if r['ma'].verfuegbarkeit not in ('teilzeit', 'zusatz')
    ]
    if vollzeit_rows:
        avg_t = sum(r['t'] for r in vollzeit_rows) / len(vollzeit_rows)
        avg_n = sum(r['n'] for r in vollzeit_rows) / len(vollzeit_rows)
        avg_we = sum(r['we'] for r in vollzeit_rows) / len(vollzeit_rows)
        max_dev_t = max(abs(r['t'] - avg_t) for r in vollzeit_rows) or 1
        max_dev_n = max(abs(r['n'] - avg_n) for r in vollzeit_rows) or 1
        max_dev_we = max(abs(r['we'] - avg_we) for r in vollzeit_rows) or 1
        for r in jahres_mitarbeiter:
            if r['ma'].verfuegbarkeit in ('teilzeit', 'zusatz'):
                r['heat_t'] = ''
                r['heat_n'] = ''
                r['heat_we'] = ''
                continue
            score_t = abs(r['t'] - avg_t) / max_dev_t
            score_n = abs(r['n'] - avg_n) / max_dev_n
            score_we = abs(r['we'] - avg_we) / max_dev_we
            hue_t = 120 - int(120 * score_t)
            hue_n = 120 - int(120 * score_n)
            hue_we = 120 - int(120 * score_we)
            r['heat_t'] = f"hsl({hue_t}, 70%, 85%)"
            r['heat_n'] = f"hsl({hue_n}, 70%, 85%)"
            r['heat_we'] = f"hsl({hue_we}, 70%, 85%)"
    else:
        for r in jahres_mitarbeiter:
            r['heat_t'] = ''
            r['heat_n'] = ''
            r['heat_we'] = ''
    jahres_mitarbeiter.sort(key=lambda x: (
        x['ma'].verfuegbarkeit in ('teilzeit', 'zusatz'),
        len(x['ma'].schichtplan_kennung or ''),
        x['ma'].schichtplan_kennung or ''
    ))
    jahresuebersicht = {
        'jahr': jahr,
        'jahres_mitarbeiter': jahres_mitarbeiter,
    }

    context = {
        'schichtplan': schichtplan,
        'kalender_daten': kalender_daten,
        'mitarbeiter_stats': stats_list,
        'schichttypen': Schichttyp.objects.filter(aktiv=True),
        'mitarbeiter_mapping': mitarbeiter_mapping,
        'can_edit': ist_schichtplaner(request.user),
        'user_is_staff': request.user.is_staff,
        'jahresuebersicht': jahresuebersicht,
    }

    return render(request, 'schichtplan/schichtplan_detail.html', context)


@login_required
def schichtplan_zur_genehmigung(request, pk):
    """Plan von Entwurf auf 'Zur Genehmigung' setzen (nur Schichtplaner, POST)."""
    schichtplan = get_object_or_404(Schichtplan, pk=pk)
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('arbeitszeit:dashboard')
    if schichtplan.status != 'entwurf':
        messages.warning(request, f"Plan hat Status „{schichtplan.get_status_display()}“. Nur Entwürfe können zur Genehmigung gegeben werden.")
        return redirect('schichtplan:detail', pk=pk)
    if request.method != 'POST':
        return redirect('schichtplan:detail', pk=pk)
    schichtplan.status = 'zur_genehmigung'
    schichtplan.save(update_fields=['status', 'aktualisiert_am'])
    messages.success(request, "✅ Schichtplan wurde zur Genehmigung weitergeleitet.")
    return redirect('schichtplan:detail', pk=pk)


@login_required
def schichtplan_veroeffentlichen(request, pk):
    """Plan durch Admin genehmigen und veröffentlichen (nur Staff, POST)."""
    schichtplan = get_object_or_404(Schichtplan, pk=pk)
    if not request.user.is_staff:
        messages.error(request, "❌ Nur Administratoren können Pläne veröffentlichen.")
        return redirect('arbeitszeit:dashboard')
    if schichtplan.status != 'zur_genehmigung':
        messages.warning(request, f"Plan hat Status „{schichtplan.get_status_display()}“. Nur Pläne „Zur Genehmigung“ können veröffentlicht werden.")
        return redirect('schichtplan:detail', pk=pk)
    if request.method != 'POST':
        return redirect('schichtplan:detail', pk=pk)
    with transaction.atomic():
        schichtplan.status = 'veroeffentlicht'
        schichtplan.save(update_fields=['status', 'aktualisiert_am'])
        # Snapshot der Schichten zum Zeitpunkt der Veröffentlichung
        SchichtplanSnapshot.objects.filter(schichtplan=schichtplan).delete()
        snapshot = SchichtplanSnapshot.objects.create(
            schichtplan=schichtplan,
            erstellt_von=request.user
        )
        snapshot_rows = [
            SchichtplanSnapshotSchicht(
                snapshot=snapshot,
                mitarbeiter_id=s.mitarbeiter_id,
                datum=s.datum,
                schichttyp_id=s.schichttyp_id
            )
            for s in schichtplan.schichten.all()
        ]
        if snapshot_rows:
            SchichtplanSnapshotSchicht.objects.bulk_create(snapshot_rows)
    messages.success(request, "✅ Schichtplan wurde genehmigt und veröffentlicht. Die Abteilung Kongos hat nun Zugriff.")
    return redirect('schichtplan:detail', pk=pk)


@login_required
def schichtplan_uebersicht_detail(request, pk):
    """
    Tabellarische Übersicht: Zeilen = Tage, Spalten = MA1–MA15 (Vollname).
    Zellen: N, T, Z, U, AG. Kongos haben Lesezugriff auf veröffentlichte Pläne.
    """
    schichtplan = get_object_or_404(Schichtplan, pk=pk)
    if not darf_schichtplan_sehen(request.user, schichtplan):
        messages.error(request, "❌ Keine Berechtigung für diesen Schichtplan.")
        return redirect('arbeitszeit:dashboard')

    # Mitarbeiter MA1–MA15, sortiert
    alle_ma = get_planbare_mitarbeiter()
    def ma_sort_key(ma):
        k = ma.schichtplan_kennung or ''
        if k.startswith('MA') and len(k) > 2:
            try:
                return int(k[2:])
            except ValueError:
                pass
        return 999
    mitarbeiter_liste = sorted(alle_ma, key=ma_sort_key)

    # Tage des Plans
    start = schichtplan.start_datum
    ende = schichtplan.ende_datum
    tage_liste = []
    d = start
    while d <= ende:
        tage_liste.append(d)
        d += timedelta(days=1)

    # NRW-Feiertage im Planzeitraum
    feiertage_set, feiertage_namen = get_configured_feiertage(start, ende, region='nrw')

    # Schichten: (ma_id, datum) -> Kürzel und Schicht-ID (für Löschen/Bearbeiten/Tauschen)
    schichten = schichtplan.schichten.select_related('mitarbeiter', 'schichttyp')
    zelle_schicht = {}
    zelle_schicht_id = {}
    zelle_schicht_ersatz = {}
    for s in schichten:
        key = (s.mitarbeiter_id, s.datum)
        zelle_schicht[key] = s.schichttyp.kuerzel
        zelle_schicht_id[key] = s.pk
        zelle_schicht_ersatz[key] = bool(s.ersatz_markierung)

    # Wünsche U/K/AG: Sets und (ma_id, datum) -> wunsch_id für Löschen per Drag
    wunsch_urlaub = Schichtwunsch.objects.filter(
        datum__gte=start, datum__lte=ende,
        mitarbeiter__in=mitarbeiter_liste,
        wunsch='urlaub'
    ).values_list('mitarbeiter_id', 'datum', 'pk')
    wunsch_gar_nichts = Schichtwunsch.objects.filter(
        datum__gte=start, datum__lte=ende,
        mitarbeiter__in=mitarbeiter_liste,
        wunsch='gar_nichts'
    ).values_list('mitarbeiter_id', 'datum', 'pk')
    urlaub_set = set((ma_id, datum) for ma_id, datum, _ in wunsch_urlaub)
    gar_nichts_set = set((ma_id, datum) for ma_id, datum, _ in wunsch_gar_nichts)
    zelle_wunsch_id = {(ma_id, datum): pk for ma_id, datum, pk in wunsch_urlaub}
    zelle_wunsch_id.update({(ma_id, datum): pk for ma_id, datum, pk in wunsch_gar_nichts})

    wunsch_krank = Schichtwunsch.objects.filter(
        datum__gte=start, datum__lte=ende,
        mitarbeiter__in=mitarbeiter_liste,
        wunsch='krank'
    ).values_list('mitarbeiter_id', 'datum', 'pk', 'ersatz_schichttyp', 'ersatz_bestaetigt')
    krank_set = set((ma_id, datum) for ma_id, datum, _, _, _ in wunsch_krank)
    ersatz_info = {}
    for ma_id, datum, pk, ersatz_typ, ersatz_bestaetigt in wunsch_krank:
        zelle_wunsch_id[(ma_id, datum)] = pk
        ersatz_info[(ma_id, datum)] = {
            'typ': ersatz_typ or '',
            'bestaetigt': bool(ersatz_bestaetigt),
        }

    wunsch_ausgleich = Schichtwunsch.objects.filter(
        datum__gte=start, datum__lte=ende,
        mitarbeiter__in=mitarbeiter_liste,
        wunsch='ausgleichstag'
    ).values_list('mitarbeiter_id', 'datum', 'pk')
    ausgleich_set = set((ma_id, datum) for ma_id, datum, _ in wunsch_ausgleich)
    for ma_id, datum, pk in wunsch_ausgleich:
        zelle_wunsch_id[(ma_id, datum)] = pk

    # Wunsch-Overlay aus Wunschperiode (alle Kategorien, als Hint)
    wunsch_overlay = {}
    wunsch_overlay_label = {}
    wunsch_map = {
        'urlaub': ('U', 'Urlaub'),
        'kein_tag_aber_nacht': ('2', 'Kein Tag, Nacht OK'),
        'keine_nacht_aber_tag': ('3', 'Keine Nacht, Tag OK'),
        'tag_bevorzugt': ('T', 'Tag bevorzugt'),
        'nacht_bevorzugt': ('N', 'Nacht bevorzugt'),
        'gar_nichts': ('X', 'Gar nichts'),
        'zusatzarbeit': ('+', 'Zusatzarbeit'),
    }
    overlay_qs = Schichtwunsch.objects.filter(
        datum__gte=start,
        datum__lte=ende,
        mitarbeiter__in=mitarbeiter_liste,
    )
    if schichtplan.wunschperiode_id:
        overlay_qs = overlay_qs.filter(periode=schichtplan.wunschperiode)
    for w in overlay_qs.select_related('mitarbeiter'):
        symb_label = wunsch_map.get(w.wunsch)
        if not symb_label:
            continue
        key = (w.mitarbeiter_id, w.datum)
        wunsch_overlay[key] = symb_label[0]
        wunsch_overlay_label[key] = symb_label[1]

    # Matrix: (ma_id, datum) -> 'N'|'T'|'Z'|'U'|'AG'|''
    matrix = {}
    for ma in mitarbeiter_liste:
        for datum in tage_liste:
            key = (ma.id, datum)
            if key in zelle_schicht:
                matrix[key] = zelle_schicht[key]
            elif key in urlaub_set:
                matrix[key] = 'U'
            elif key in gar_nichts_set:
                matrix[key] = 'X'
            elif key in krank_set:
                matrix[key] = 'K'
            elif key in ausgleich_set:
                matrix[key] = 'AG'
            else:
                matrix[key] = ''

    # Zeilen = Tage: pro Zeile (datum, ist_wochenende, feiertag_name, [Zellen pro MA])
    # Jede Zelle: {'value': ..., 'schicht_id': int|None, 'wunsch_id': int|None, 'ma_id': int}
    zeilen = []
    for datum in tage_liste:
        ist_wochenende = datum.weekday() >= 5  # Sa=5, So=6
        feiertag_name = feiertage_namen.get(datum, '')
        zellen = []
        t_count = 0
        n_count = 0
        for ma in mitarbeiter_liste:
            key = (ma.id, datum)
            val = matrix.get(key, '')
            if val == 'T' and ma.zaehlt_zur_tagbesetzung:
                t_count += 1
            elif val == 'N':
                n_count += 1
            sid = zelle_schicht_id.get(key) if val in ('T', 'N', 'Z') else None
            wid = zelle_wunsch_id.get(key) if val in ('U', 'K', 'AG') else None
            ersatz_markierung = zelle_schicht_ersatz.get(key, False) if val in ('T', 'N', 'Z') else False
            ersatz_typ = ''
            ersatz_bestaetigt = False
            if val == 'K':
                info = ersatz_info.get(key, {})
                ersatz_typ = info.get('typ', '')
                ersatz_bestaetigt = info.get('bestaetigt', False)
            overlay_symbol = wunsch_overlay.get(key, '')
            overlay_label = wunsch_overlay_label.get(key, '')
            zellen.append({
                'value': val,
                'schicht_id': sid,
                'wunsch_id': wid,
                'ma_id': ma.id,
                'ersatz_markierung': ersatz_markierung,
                'ersatz_typ': ersatz_typ,
                'ersatz_bestaetigt': ersatz_bestaetigt,
                'wunsch_overlay': overlay_symbol,
                'wunsch_overlay_label': overlay_label,
            })
        zeilen.append({
            'datum': datum,
            'ist_wochenende': ist_wochenende,
            'feiertag_name': feiertag_name,
            'zellen': zellen,
            't_count': t_count,
            'n_count': n_count,
        })

    ersatz_vorschlaege = []
    if schichtplan.status == 'veroeffentlicht':
        wunsch_block_set = urlaub_set | krank_set | ausgleich_set
        kernteam = list(alle_ma.filter(kategorie='kern'))
        kernteam_sorted = sorted(kernteam, key=ma_sort_key)

        def _hat_schicht(ma_id, datum, kuerzel_set):
            return zelle_schicht.get((ma_id, datum)) in kuerzel_set

        def _ist_blockiert(ma_id, datum):
            return (ma_id, datum) in wunsch_block_set

        for ma_id, datum in krank_set:
            krank_ma = next((m for m in mitarbeiter_liste if m.id == ma_id), None)
            if not krank_ma:
                continue
            info = ersatz_info.get((ma_id, datum), {})
            if not info.get('typ'):
                continue
            wunsch_id = zelle_wunsch_id.get((ma_id, datum))
            tag_kandidaten = []
            nacht_kandidaten = []
            next_day = datum + timedelta(days=1)
            prev_day = datum - timedelta(days=1)
            for kandidat in kernteam_sorted:
                if kandidat.id == ma_id:
                    continue
                if _ist_blockiert(kandidat.id, datum):
                    continue
                if not _hat_schicht(kandidat.id, datum, {'T', 'N'}) and not _hat_schicht(kandidat.id, prev_day, {'N'}):
                    tag_kandidaten.append(kandidat)
                if not _hat_schicht(kandidat.id, datum, {'N'}) and not _hat_schicht(kandidat.id, next_day, {'T'}):
                    nacht_kandidaten.append(kandidat)
            ersatz_vorschlaege.append({
                'datum': datum,
                'krank_ma': krank_ma,
                'wunsch_id': wunsch_id,
                'tag_kandidaten': tag_kandidaten,
                'nacht_kandidaten': nacht_kandidaten,
            })

    # Schichttypen T, N, Z für das „Schicht zuweisen“-Panel (Drag-Quelle)
    schichttypen_menu = list(Schichttyp.objects.filter(kuerzel__in=['T', 'N', 'Z'], aktiv=True).order_by('kuerzel'))

    # NEU: Feiertage einmalig berechnen
    feiertage_set, _ = get_configured_feiertage(start, ende, region='nrw')

    # Soll- und Ist-Stunden pro MA wie auf Plan-Detailseite (für Abschlusszeile pro Spalte)
    stunden_defaults = {'T': 12.25, 'N': 12.25, 'Z': 8.0}
    urlaub_krank_pro_ma = defaultdict(list)
    for ma_id, datum in urlaub_set | krank_set:
        urlaub_krank_pro_ma[ma_id].append(datum)
    mitarbeiter_stunden = []  # Liste (soll, ist, diff) in gleicher Reihenfolge wie mitarbeiter_liste
    for ma in mitarbeiter_liste:
        ma_schichten = [s for s in schichten if s.mitarbeiter_id == ma.id]
        ist_stunden = 0.0
        for s in ma_schichten:
            kuerzel = s.schichttyp.kuerzel
            if ma.schichtplan_kennung == 'MA7' and kuerzel == 'Z':
                stunden = 12.25
            elif kuerzel == 'Z':
                vereinbarung = ma.get_aktuelle_vereinbarung(s.datum)
                if vereinbarung and vereinbarung.wochenstunden:
                    stunden = float(vereinbarung.wochenstunden) / 5.0
                else:
                    stunden = float(s.schichttyp.arbeitszeit_stunden) if s.schichttyp.arbeitszeit_stunden else stunden_defaults.get(kuerzel, 0)
            else:
                stunden = float(s.schichttyp.arbeitszeit_stunden) if s.schichttyp.arbeitszeit_stunden else stunden_defaults.get(kuerzel, 0)
            ist_stunden += stunden
        ist_stunden += _ist_stunden_urlaub_krank(ma, urlaub_krank_pro_ma.get(ma.id, []), feiertage_set)
        try:
            soll_stunden = float(ma.get_soll_stunden_monat(start.year, start.month))
        except Exception:
            soll_stunden = 160.0
        diff = ist_stunden - soll_stunden
        mitarbeiter_stunden.append({'soll': soll_stunden, 'ist': ist_stunden, 'diff': diff})

    protokoll = (
        SchichtplanAenderung.objects.filter(schichtplan=schichtplan)
        .select_related('user')
        .order_by('-zeit')[:50]
    )
    letzte_aenderung = protokoll.first() if protokoll else None
    can_edit = darf_plan_bearbeiten(request.user, schichtplan)
    kann_undo = can_edit and letzte_aenderung and not letzte_aenderung.zurueckgenommen

    context = {
        'schichtplan': schichtplan,
        'mitarbeiter_liste': mitarbeiter_liste,
        'tage_liste': tage_liste,
        'zeilen': zeilen,
        'feiertage_namen': feiertage_namen,
        'schichttypen_menu': schichttypen_menu,
        'mitarbeiter_stunden': mitarbeiter_stunden,
        'can_edit': can_edit,
        'protokoll': protokoll,
        'kann_undo': kann_undo,
        'ersatz_vorschlaege': ersatz_vorschlaege,
    }
    return render(request, 'schichtplan/schichtplan_uebersicht_detail.html', context)


@login_required
def schichtplan_rueckgaengig(request, pk):
    """Letzte Änderung an diesem Schichtplan rückgängig machen (Schichtplaner und Kongos bei veröff. Plan)."""
    schichtplan = get_object_or_404(Schichtplan, pk=pk)
    if not darf_plan_bearbeiten(request.user, schichtplan):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('schichtplan:dashboard')
    if request.method != 'POST':
        return redirect('schichtplan:uebersicht_detail', pk=pk)

    letzte = (
        SchichtplanAenderung.objects.filter(schichtplan=schichtplan, zurueckgenommen=False)
        .order_by('-zeit')
        .first()
    )
    if not letzte:
        messages.warning(request, "Keine Änderung zum Rückgängigmachen vorhanden.")
        return redirect('schichtplan:uebersicht_detail', pk=pk)

    from datetime import datetime as dt
    ud = letzte.undo_daten or {}

    try:
        with transaction.atomic():
            if letzte.aktion == 'angelegt':
                schicht_id = ud.get('schicht_id')
                if schicht_id:
                    Schicht.objects.filter(pk=schicht_id, schichtplan=schichtplan).delete()
            elif letzte.aktion == 'geloescht':
                ma_id = ud.get('mitarbeiter_id')
                datum_str = ud.get('datum')
                typ_id = ud.get('schichttyp_id')
                if ma_id and datum_str and typ_id:
                    datum = dt.strptime(datum_str, '%Y-%m-%d').date()
                    ma = get_planbare_mitarbeiter().filter(pk=ma_id).first()
                    typ = Schichttyp.objects.filter(pk=typ_id).first()
                    if ma and typ and schichtplan.start_datum <= datum <= schichtplan.ende_datum:
                        Schicht.objects.get_or_create(
                            schichtplan=schichtplan,
                            mitarbeiter_id=ma_id,
                            datum=datum,
                            defaults={'schichttyp_id': typ_id},
                        )
            elif letzte.aktion == 'getauscht':
                s1_id = ud.get('schicht1_id')
                s2_id = ud.get('schicht2_id')
                if s1_id and s2_id:
                    s1 = Schicht.objects.filter(pk=s1_id, schichtplan=schichtplan).first()
                    s2 = Schicht.objects.filter(pk=s2_id, schichtplan=schichtplan).first()
                    if s1 and s2:
                        typ1, typ2 = s1.schichttyp_id, s2.schichttyp_id
                        s1.schichttyp_id = typ2
                        s2.schichttyp_id = typ1
                        s1.save(update_fields=['schichttyp_id'])
                        s2.save(update_fields=['schichttyp_id'])
            letzte.zurueckgenommen = True
            letzte.save(update_fields=['zurueckgenommen'])
        messages.success(request, "✅ Letzte Änderung wurde rückgängig gemacht.")
    except Exception as e:
        messages.error(request, f"❌ Rückgängig fehlgeschlagen: {e}")
    return redirect('schichtplan:uebersicht_detail', pk=pk)


@login_required
def schichtplan_export_excel(request, pk):
    """Schichtplan (Übersicht: Tage × MA) als Excel exportieren. Kongos dürfen veröffentlichte Pläne exportieren."""
    schichtplan = get_object_or_404(Schichtplan, pk=pk)
    if not darf_schichtplan_sehen(request.user, schichtplan):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('schichtplan:dashboard')

    alle_ma = get_planbare_mitarbeiter()
    def ma_sort_key(ma):
        k = ma.schichtplan_kennung or ''
        if k.startswith('MA') and len(k) > 2:
            try:
                return int(k[2:])
            except ValueError:
                pass
        return 999
    mitarbeiter_liste = sorted(alle_ma, key=ma_sort_key)
    start = schichtplan.start_datum
    ende = schichtplan.ende_datum
    tage_liste = []
    d = start
    while d <= ende:
        tage_liste.append(d)
        d += timedelta(days=1)

    schichten = schichtplan.schichten.select_related('mitarbeiter', 'schichttyp')
    zelle_schicht = {}
    for s in schichten:
        zelle_schicht[(s.mitarbeiter_id, s.datum)] = s.schichttyp.kuerzel

    urlaub_wuensche = Schichtwunsch.objects.filter(
        datum__gte=start, datum__lte=ende, mitarbeiter__in=mitarbeiter_liste,
        wunsch='urlaub'
    ).values_list('mitarbeiter_id', 'datum', 'pk')
    gar_nichts_wuensche = Schichtwunsch.objects.filter(
        datum__gte=start, datum__lte=ende, mitarbeiter__in=mitarbeiter_liste,
        wunsch='gar_nichts'
    ).values_list('mitarbeiter_id', 'datum', 'pk')
    urlaub_set = set((ma_id, datum) for ma_id, datum, _ in urlaub_wuensche)
    gar_nichts_set = set((ma_id, datum) for ma_id, datum, _ in gar_nichts_wuensche)
    krank_set = set(Schichtwunsch.objects.filter(
        datum__gte=start, datum__lte=ende, mitarbeiter__in=mitarbeiter_liste,
        wunsch='krank'
    ).values_list('mitarbeiter_id', 'datum'))
    ausgleich_set = set(Schichtwunsch.objects.filter(
        datum__gte=start, datum__lte=ende, mitarbeiter__in=mitarbeiter_liste,
        wunsch='ausgleichstag'
    ).values_list('mitarbeiter_id', 'datum'))
    urlaub_krank_pro_ma = defaultdict(list)
    for ma_id, datum in urlaub_set | krank_set:
        urlaub_krank_pro_ma[ma_id].append(datum)

    matrix = {}
    for ma in mitarbeiter_liste:
        for datum in tage_liste:
            key = (ma.id, datum)
            if key in zelle_schicht:
                matrix[key] = zelle_schicht[key]
            elif key in urlaub_set:
                matrix[key] = 'U'
            elif key in gar_nichts_set:
                matrix[key] = 'X'
            elif key in krank_set:
                matrix[key] = 'K'
            elif key in ausgleich_set:
                matrix[key] = 'AG'
            else:
                matrix[key] = ''

    stunden_defaults = {'T': 12.25, 'N': 12.25, 'Z': 8.0}
    # NEU: Feiertage für Ist-Stunden-Berechnung
    feiertage_set, _ = get_configured_feiertage(start, ende, region='nrw')
    mitarbeiter_stunden = []
    for ma in mitarbeiter_liste:
        ma_schichten = [s for s in schichten if s.mitarbeiter_id == ma.id]
        ist_stunden = 0.0
        for s in ma_schichten:
            kuerzel = s.schichttyp.kuerzel
            if ma.schichtplan_kennung == 'MA7' and kuerzel == 'Z':
                stunden = 12.25
            elif kuerzel == 'Z':
                vereinbarung = ma.get_aktuelle_vereinbarung(s.datum)
                stunden = (float(vereinbarung.wochenstunden) / 5.0) if vereinbarung and vereinbarung.wochenstunden else (float(s.schichttyp.arbeitszeit_stunden) if s.schichttyp.arbeitszeit_stunden else stunden_defaults.get(kuerzel, 0))
            else:
                stunden = float(s.schichttyp.arbeitszeit_stunden) if s.schichttyp.arbeitszeit_stunden else stunden_defaults.get(kuerzel, 0)
            ist_stunden += stunden
        ist_stunden += _ist_stunden_urlaub_krank(ma, urlaub_krank_pro_ma.get(ma.id, []), feiertage_set)
        try:
            soll_stunden = float(ma.get_soll_stunden_monat(start.year, start.month))
        except Exception:
            soll_stunden = 160.0
        mitarbeiter_stunden.append({'soll': soll_stunden, 'ist': ist_stunden, 'diff': ist_stunden - soll_stunden})

    # NRW-Feiertage und deutsche Wochentage für Markierungen
    _, feiertage_namen = get_configured_feiertage(start, ende, region='nrw')
    wochentage_de = ['Mo', 'Di', 'Mi', 'Do', 'Fr', 'Sa', 'So']

    # Farben (openpyxl: RRGGBB ohne #)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    fill_wochenende = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")   # hellblau
    fill_feiertag = PatternFill(start_color="FFE4CC", end_color="FFE4CC", fill_type="solid")    # hellorange
    fill_t = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")           # gelb (Tagschicht)
    fill_n = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")            # blau (Nacht)
    fill_z = PatternFill(start_color="9EEAF9", end_color="9EEAF9", fill_type="solid")            # hellblau (Zusatz)
    fill_u = PatternFill(start_color="ADB5BD", end_color="ADB5BD", fill_type="solid")           # grau (Urlaub)
    fill_k = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")           # rot (Krank)
    fill_ag = PatternFill(start_color="212529", end_color="212529", fill_type="solid")          # dunkel (Zeitausgleich)
    font_hell = Font(color="FFFFFF")
    zellen_farben = {'T': (fill_t, None), 'N': (fill_n, font_hell), 'Z': (fill_z, None), 'U': (fill_u, font_hell), 'K': (fill_k, font_hell), 'AG': (fill_ag, font_hell)}

    wb = Workbook()
    ws = wb.active
    ws.title = "Schichtplan"[:31]

    # Header-Zeile
    headers = ['Tag', 'Datum'] + [ma.schichtplan_kennung for ma in mitarbeiter_liste]
    for col, header in enumerate(headers, 1):
        c = ws.cell(1, col, header)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')

    # Datenzeilen mit deutschen Wochentagen und farblichen Markierungen
    for row_idx, datum in enumerate(tage_liste, 2):
        ist_wochenende = datum.weekday() >= 5
        ist_feiertag = datum in feiertage_namen
        zeilen_fill = fill_feiertag if ist_feiertag else (fill_wochenende if ist_wochenende else None)

        tag_name = wochentage_de[datum.weekday()]
        if ist_feiertag:
            tag_name = f"{tag_name} ({feiertage_namen[datum]})"
        c1 = ws.cell(row_idx, 1, tag_name)
        c2 = ws.cell(row_idx, 2, datum.strftime('%d.%m.%Y'))
        if zeilen_fill:
            c1.fill = zeilen_fill
            c2.fill = zeilen_fill
        for col_idx, ma in enumerate(mitarbeiter_liste, 3):
            val = matrix.get((ma.id, datum), '')
            cell = ws.cell(row_idx, col_idx, val if val else '')
            if val and val in zellen_farben:
                fill, font = zellen_farben[val]
                cell.fill = fill
                if font:
                    cell.font = font
            elif zeilen_fill:
                cell.fill = zeilen_fill

    # Abschlussstrich (dicker Rahmen) + Summenzeilen
    num_cols = 2 + len(mitarbeiter_liste)
    thin = Side(style='thin')
    thick_top = Side(style='medium')
    abschluss_border = Border(
        left=thin, right=thin, top=thick_top, bottom=thin
    )
    footer_row = len(tage_liste) + 2
    for col in range(1, num_cols + 1):
        c = ws.cell(footer_row, col)
        if col == 1:
            c.value = 'Soll (h)'
        elif col == 2:
            c.value = ''
        else:
            c.value = round(mitarbeiter_stunden[col - 3]['soll'], 1)
        c.border = abschluss_border
    footer_row += 1
    for col in range(1, num_cols + 1):
        c = ws.cell(footer_row, col)
        if col == 1:
            c.value = 'Ist (h)'
        elif col == 2:
            c.value = ''
        else:
            c.value = round(mitarbeiter_stunden[col - 3]['ist'], 1)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    footer_row += 1
    for col in range(1, num_cols + 1):
        c = ws.cell(footer_row, col)
        if col == 1:
            c.value = 'Differenz (h)'
        elif col == 2:
            c.value = ''
        else:
            c.value = round(mitarbeiter_stunden[col - 3]['diff'], 1)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    for col in range(3, len(mitarbeiter_liste) + 3):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 6

    safe_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in schichtplan.name)[:50]
    filename = f"Schichtplan_{safe_name}_{start.year}-{start.month:02d}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def schichtplan_loeschen(request, pk):
    schichtplan = get_object_or_404(Schichtplan, pk=pk)
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('schichtplan:dashboard')
    if schichtplan.status != 'entwurf':
        messages.error(request, "❌ Nur Entwuerfe koennen geloescht werden.")
        return redirect('schichtplan:dashboard')
    if request.method == 'POST':
        name = schichtplan.name
        schichtplan.delete()
        messages.success(request, f"✅ Entwurf '{name}' wurde geloescht.")
        return redirect('schichtplan:dashboard')
    return render(request, 'schichtplan/schichtplan_loeschen_confirm.html', {
        'schichtplan': schichtplan,
    })


@login_required
def schicht_anlegen(request, schichtplan_pk):
    """Schicht oder Markierung (U/K/AG) per Drag&Drop anlegen. Schichtplaner und Kongos (bei veröff. Plan) – alle Änderungen protokolliert."""
    schichtplan = get_object_or_404(Schichtplan, pk=schichtplan_pk)
    if not darf_plan_bearbeiten(request.user, schichtplan):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('schichtplan:dashboard')
    if request.method != 'POST':
        return redirect('schichtplan:uebersicht_detail', pk=schichtplan_pk)
    ma_id = request.POST.get('mitarbeiter_id')
    datum_str = request.POST.get('datum')
    kuerzel = (request.POST.get('schichttyp_kuerzel') or '').strip().upper()
    if not ma_id or not datum_str or not kuerzel:
        messages.error(request, "❌ Ungültige Angaben (Mitarbeiter, Datum oder Typ fehlt).")
        return redirect('schichtplan:uebersicht_detail', pk=schichtplan_pk)
    try:
        datum = datetime.strptime(datum_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "❌ Ungültiges Datum.")
        return redirect('schichtplan:uebersicht_detail', pk=schichtplan_pk)
    if datum < schichtplan.start_datum or datum > schichtplan.ende_datum:
        messages.error(request, "❌ Datum liegt außerhalb des Planzeitraums.")
        return redirect('schichtplan:uebersicht_detail', pk=schichtplan_pk)
    ma = get_planbare_mitarbeiter().filter(pk=ma_id).first()
    if not ma:
        messages.error(request, "❌ Mitarbeiter nicht gefunden.")
        return redirect('schichtplan:uebersicht_detail', pk=schichtplan_pk)

    # U = Urlaub, K = Krank, AG = Z-AG Zeitausgleich → Schichtwunsch anlegen/aktualisieren
    if kuerzel in ('U', 'K', 'AG'):
        wunsch_map = {'U': 'urlaub', 'K': 'krank', 'AG': 'ausgleichstag'}
        wunsch = wunsch_map[kuerzel]
        bestehende_schicht = Schicht.objects.filter(
            schichtplan=schichtplan,
            mitarbeiter_id=ma_id,
            datum=datum
        ).select_related('schichttyp').first()
        ersatz_typ = ''
        if kuerzel == 'K' and bestehende_schicht and bestehende_schicht.schichttyp:
            if bestehende_schicht.schichttyp.kuerzel in ('T', 'N'):
                ersatz_typ = bestehende_schicht.schichttyp.kuerzel
        # Bestehende Schicht an diesem Tag entfernen (Plan zeigt dann U/K/AG)
        Schicht.objects.filter(schichtplan=schichtplan, mitarbeiter_id=ma_id, datum=datum).delete()
        wunsch_obj, created = Schichtwunsch.objects.update_or_create(
            mitarbeiter_id=int(ma_id),
            datum=datum,
            defaults={
                'wunsch': wunsch,
                'periode': schichtplan.wunschperiode,
                'genehmigt': True,
                'ersatz_schichttyp': ersatz_typ if kuerzel == 'K' else '',
                'ersatz_bestaetigt': False,
                'ersatz_mitarbeiter': None,
            }
        )
        return redirect('schichtplan:uebersicht_detail', pk=schichtplan_pk)

    # T, N, Z → Schicht anlegen
    if kuerzel not in ('T', 'N', 'Z'):
        messages.error(request, "❌ Ungültiger Schichttyp.")
        return redirect('schichtplan:uebersicht_detail', pk=schichtplan_pk)
    schichttyp = Schichttyp.objects.filter(kuerzel=kuerzel, aktiv=True).first()
    if not schichttyp:
        messages.error(request, f"❌ Schichttyp '{kuerzel}' nicht gefunden.")
        return redirect('schichtplan:uebersicht_detail', pk=schichtplan_pk)
    # Optional: bestehenden Wunsch (Urlaub/Krank/AG) für diesen Tag entfernen, damit Schicht sichtbar ist
    Schichtwunsch.objects.filter(mitarbeiter_id=ma_id, datum=datum, wunsch__in=['urlaub', 'gar_nichts', 'krank', 'ausgleichstag']).delete()
    if Schicht.objects.filter(schichtplan=schichtplan, mitarbeiter_id=ma_id, datum=datum).exists():
        messages.error(request, "❌ An diesem Tag hat die Person bereits eine Schicht.")
        return redirect('schichtplan:uebersicht_detail', pk=schichtplan_pk)
    schicht = Schicht.objects.create(schichtplan=schichtplan, mitarbeiter=ma, datum=datum, schichttyp=schichttyp)
    SchichtplanAenderung.objects.create(
        schichtplan=schichtplan,
        user=request.user,
        aktion='angelegt',
        beschreibung=f"{ma.schichtplan_kennung} {kuerzel} am {datum.strftime('%d.%m.%Y')} angelegt",
        undo_daten={'schicht_id': schicht.pk},
    )
    return redirect('schichtplan:uebersicht_detail', pk=schichtplan_pk)


@login_required
def ersatz_bestaetigen(request, pk):
    if request.method != 'POST':
        return redirect('schichtplan:uebersicht_detail', pk=pk)
    schichtplan = get_object_or_404(Schichtplan, pk=pk)
    if schichtplan.status != 'veroeffentlicht' or not darf_plan_bearbeiten(request.user, schichtplan):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('schichtplan:uebersicht_detail', pk=pk)
    wunsch_id = request.POST.get('wunsch_id')
    ersatz_ma_id = request.POST.get('ersatz_ma_id')
    if not wunsch_id:
        return redirect('schichtplan:uebersicht_detail', pk=pk)
    wunsch = get_object_or_404(Schichtwunsch, pk=wunsch_id, wunsch='krank')
    if wunsch.datum < schichtplan.start_datum or wunsch.datum > schichtplan.ende_datum:
        return redirect('schichtplan:uebersicht_detail', pk=pk)
    if not wunsch.ersatz_schichttyp:
        return redirect('schichtplan:uebersicht_detail', pk=pk)
    if not ersatz_ma_id:
        return redirect('schichtplan:uebersicht_detail', pk=pk)
    ersatz_ma = get_planbare_mitarbeiter().filter(pk=ersatz_ma_id).first()
    if not ersatz_ma:
        return redirect('schichtplan:uebersicht_detail', pk=pk)
    bestehende_schicht = Schicht.objects.filter(
        schichtplan=schichtplan,
        mitarbeiter_id=ersatz_ma.id,
        datum=wunsch.datum
    ).select_related('schichttyp').first()
    if bestehende_schicht:
        if bestehende_schicht.schichttyp and bestehende_schicht.schichttyp.kuerzel == 'Z':
            bestehende_schicht.delete()
        else:
            messages.error(request, "❌ Ersatz hat bereits eine Schicht an diesem Tag.")
            return redirect('schichtplan:uebersicht_detail', pk=pk)
    if Schichtwunsch.objects.filter(
        mitarbeiter_id=ersatz_ma.id,
        datum=wunsch.datum,
        wunsch__in=['urlaub', 'gar_nichts', 'krank', 'ausgleichstag']
    ).exists():
        messages.error(request, "❌ Ersatz hat eine Markierung (U/K/AG) an diesem Tag.")
        return redirect('schichtplan:uebersicht_detail', pk=pk)
    schichttyp = Schichttyp.objects.filter(kuerzel=wunsch.ersatz_schichttyp, aktiv=True).first()
    if not schichttyp:
        return redirect('schichtplan:uebersicht_detail', pk=pk)
    Schicht.objects.create(
        schichtplan=schichtplan,
        mitarbeiter=ersatz_ma,
        datum=wunsch.datum,
        schichttyp=schichttyp,
        ersatz_markierung=True
    )
    wunsch.ersatz_bestaetigt = True
    wunsch.ersatz_mitarbeiter = ersatz_ma
    wunsch.save(update_fields=['ersatz_bestaetigt', 'ersatz_mitarbeiter'])
    return redirect('schichtplan:uebersicht_detail', pk=pk)


@login_required
def schicht_zuweisen(request, schichtplan_pk):
    """Schicht manuell zuweisen (Formular-Seite)"""
    schichtplan = get_object_or_404(Schichtplan, pk=schichtplan_pk)
    
    if not ist_schichtplaner(request.user):
        return JsonResponse({'error': 'Keine Berechtigung'}, status=403)
    
    if request.method == 'POST':
        form = SchichtForm(request.POST)
        
        if form.is_valid():
            schicht = form.save(commit=False)
            schicht.schichtplan = schichtplan
            
            try:
                schicht.save()
                messages.success(request, "✅ Schicht erfolgreich zugewiesen!")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'schicht_id': schicht.pk,
                    })
                
                if request.GET.get('next') == 'uebersicht_detail':
                    return redirect('schichtplan:uebersicht_detail', pk=schichtplan.pk)
                return redirect('schichtplan:detail', pk=schichtplan.pk)
                
            except Exception as e:
                messages.error(request, f"❌ Fehler beim Speichern: {e}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'error': str(e)}, status=400)
        
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': form.errors}, status=400)
    
    else:
        form = SchichtForm()
    
    context = {
        'schichtplan': schichtplan,
        'form': form,
    }
    
    return render(request, 'schichtplan/schicht_zuweisen.html', context)


@login_required
def schicht_loeschen(request, pk):
    """Schicht löschen. Schichtplaner und Kongos (bei veröff. Plan) – protokolliert."""
    schicht = get_object_or_404(Schicht, pk=pk)
    
    if not darf_plan_bearbeiten(request.user, schicht.schichtplan):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('schichtplan:dashboard')
    
    schichtplan_pk = schicht.schichtplan.pk
    next_url = request.GET.get('next', '').strip()
    
    if request.method == 'POST':
        undo_daten = {
            'mitarbeiter_id': schicht.mitarbeiter_id,
            'datum': schicht.datum.isoformat(),
            'schichttyp_id': schicht.schichttyp_id,
        }
        schichtplan_obj = schicht.schichtplan
        ma_kennung = schicht.mitarbeiter.schichtplan_kennung or ''
        kuerzel = schicht.schichttyp.kuerzel
        datum_str = schicht.datum.strftime('%d.%m.%Y')
        schicht.delete()
        SchichtplanAenderung.objects.create(
            schichtplan=schichtplan_obj,
            user=request.user,
            aktion='geloescht',
            beschreibung=f"{ma_kennung} {kuerzel} am {datum_str} gelöscht",
            undo_daten=undo_daten,
        )
        messages.success(request, "✅ Schicht wurde gelöscht.")
        if next_url == 'uebersicht_detail':
            return redirect('schichtplan:uebersicht_detail', pk=schichtplan_pk)
        return redirect('schichtplan:detail', pk=schichtplan_pk)
    
    context = {
        'schicht': schicht,
        'schichtplan_pk': schichtplan_pk,
        'next_url': next_url,
    }
    
    return render(request, 'schichtplan/schicht_loeschen_confirm.html', context)


@login_required
def schicht_bearbeiten(request, pk):
    """Schicht bearbeiten: Schichttyp und/oder Datum/Mitarbeiter ändern. Schichtplaner und Kongos – protokolliert."""
    schicht = get_object_or_404(Schicht, pk=pk)
    schichtplan = schicht.schichtplan
    if not darf_plan_bearbeiten(request.user, schichtplan):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('schichtplan:dashboard')
    next_uebersicht = request.GET.get('next') == 'uebersicht_detail'

    if request.method == 'POST':
        form = SchichtForm(request.POST, instance=schicht)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "✅ Schicht wurde aktualisiert.")
                if next_uebersicht:
                    return redirect('schichtplan:uebersicht_detail', pk=schichtplan.pk)
                return redirect('schichtplan:detail', pk=schichtplan.pk)
            except Exception as e:
                messages.error(request, f"❌ Fehler: {e}")
    else:
        form = SchichtForm(instance=schicht)

    form.fields['mitarbeiter'].queryset = get_planbare_mitarbeiter()
    form.fields['datum'].widget.attrs['min'] = schichtplan.start_datum.isoformat()
    form.fields['datum'].widget.attrs['max'] = schichtplan.ende_datum.isoformat()

    context = {
        'schicht': schicht,
        'schichtplan': schichtplan,
        'form': form,
        'next_uebersicht': next_uebersicht,
    }
    return render(request, 'schichtplan/schicht_bearbeiten.html', context)


@login_required
def schicht_tauschen(request, pk):
    """Zwei Schichten tauschen. Schichtplaner und Kongos (bei veröff. Plan) – protokolliert."""
    schicht1 = get_object_or_404(Schicht, pk=pk)
    schichtplan = schicht1.schichtplan
    if not darf_plan_bearbeiten(request.user, schichtplan):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('schichtplan:dashboard')
    next_uebersicht = request.GET.get('next') == 'uebersicht_detail'

    # Alle anderen Schichten im selben Plan (für Dropdown)
    andere_schichten = Schicht.objects.filter(schichtplan=schichtplan).exclude(pk=pk).select_related(
        'mitarbeiter', 'schichttyp'
    ).order_by('datum', 'mitarbeiter__schichtplan_kennung')

    if request.method == 'POST':
        schicht2_id = request.POST.get('schicht2_id')
        if not schicht2_id:
            messages.error(request, "❌ Bitte eine zweite Schicht auswählen.")
        else:
            schicht2 = Schicht.objects.filter(pk=schicht2_id, schichtplan=schichtplan).first()
            if not schicht2:
                messages.error(request, "❌ Ungültige Schicht.")
            else:
                # Tausch: Schichttypen der beiden Schichten vertauschen
                with transaction.atomic():
                    typ1, typ2 = schicht1.schichttyp_id, schicht2.schichttyp_id
                    schicht1.schichttyp_id = typ2
                    schicht2.schichttyp_id = typ1
                    schicht1.save(update_fields=['schichttyp_id'])
                    schicht2.save(update_fields=['schichttyp_id'])
                    SchichtplanAenderung.objects.create(
                        schichtplan=schichtplan,
                        user=request.user,
                        aktion='getauscht',
                        beschreibung=f"{schicht1.mitarbeiter.schichtplan_kennung} {schicht1.schichttyp.kuerzel} ↔ {schicht2.mitarbeiter.schichtplan_kennung} {schicht2.schichttyp.kuerzel} am {schicht1.datum.strftime('%d.%m.')} getauscht",
                        undo_daten={'schicht1_id': schicht1.pk, 'schicht2_id': schicht2.pk},
                    )
                messages.success(request, "✅ Schichten wurden getauscht.")
                if next_uebersicht:
                    return redirect('schichtplan:uebersicht_detail', pk=schichtplan.pk)
                return redirect('schichtplan:detail', pk=schichtplan.pk)

    context = {
        'schicht1': schicht1,
        'schichtplan': schichtplan,
        'andere_schichten': andere_schichten,
        'next_uebersicht': next_uebersicht,
    }
    return render(request, 'schichtplan/schicht_tauschen.html', context)


@login_required
def wuensche_genehmigen(request, periode_id):
    """Schichtplaner genehmigt Wünsche"""
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('schichtplan:dashboard')
    
    periode = get_object_or_404(SchichtwunschPeriode, pk=periode_id)
    
    context = {
        'periode': periode,
    }
    
    return render(request, 'schichtplan/wuensche_genehmigen.html', context)

@login_required
def wunsch_perioden_liste(request):
    """Liste aller Wunschperioden für Mitarbeiter"""
    if not hasattr(request.user, 'mitarbeiter'):
        messages.error(request, "❌ Kein Mitarbeiter-Profil gefunden.")
        return redirect('arbeitszeit:dashboard')
    
    mitarbeiter = request.user.mitarbeiter
    gueltige_kennungen = [f'MA{i}' for i in range(1, 16)]
    
    # Debug
    print(f"User: {request.user.username}, Kennung: {mitarbeiter.schichtplan_kennung}")
    
    if not (ist_schichtplaner(request.user) or mitarbeiter.schichtplan_kennung in gueltige_kennungen):
        messages.error(
            request, 
            f"❌ Nur MA1-MA15 können Wünsche eintragen. Ihre Kennung: {mitarbeiter.schichtplan_kennung}"
        )
        return redirect('arbeitszeit:dashboard')
    
    # Alle Perioden laden
    perioden = SchichtwunschPeriode.objects.all().order_by('-fuer_monat')
    
    # Für jede Periode: Eigene Wünsche zählen
    for periode in perioden:
        periode.eigene_wuensche_count = periode.schichtwunsch_set.filter(
            mitarbeiter=mitarbeiter
        ).count()
    
    context = {
        'perioden': perioden,
        'mitarbeiter': mitarbeiter,
        'ist_planer': ist_schichtplaner(request.user),
    }
    
    return render(request, 'schichtplan/wunsch_perioden_liste.html', context)


@login_required
def wunsch_eingeben(request, periode_id):
    """
    Formular zum Eintragen/Bearbeiten von Wünschen.
    Erlaubt Schichtplanern, Wünsche für beliebige MA einzutragen.
    NEU: Unterstützt Von-Bis Eingabe für Urlaub.
    """
    
    # 1. Grund-Berechtigung: Ist es ein Schichtplaner?
    is_planer = ist_schichtplaner(request.user)
    
    # 2. Ziel-Mitarbeiter bestimmen
    # Schichtplaner können eine mitarbeiter_id via GET/POST übergeben
    target_ma_id = request.GET.get('mitarbeiter_id') or request.POST.get('mitarbeiter_id')
    
    if is_planer and target_ma_id:
        # Admin-Modus: Nimm den MA aus dem Parameter
        mitarbeiter = get_object_or_404(Mitarbeiter, pk=target_ma_id)
    else:
        # Normaler Modus oder Planer schreibt für sich selbst:
        if not hasattr(request.user, 'mitarbeiter'):
            messages.error(request, "❌ Kein Mitarbeiter-Profil gefunden.")
            return redirect('arbeitszeit:dashboard')
        
        mitarbeiter = request.user.mitarbeiter
        
        # Prüfung für normale MA (MA1-MA15)
        gueltige_kennungen = [f'MA{i}' for i in range(1, 16)]
        if not is_planer and mitarbeiter.schichtplan_kennung not in gueltige_kennungen:
            messages.error(request, "❌ Nur MA1-MA15 können Wünsche eintragen.")
            return redirect('arbeitszeit:dashboard')
    
    # Periode laden
    periode = get_object_or_404(SchichtwunschPeriode, pk=periode_id)
    
    # Status-Check
    if not periode.ist_offen:
        messages.warning(request, f"⚠️ Wunschperiode '{periode.name}' ist geschlossen.")
        return redirect('schichtplan:wunschperioden_liste')
    
    # Datum parsen
    datum_str = request.GET.get('datum')
    if not datum_str:
        messages.error(request, "❌ Kein Datum angegeben.")
        return redirect('schichtplan:wunsch_kalender', periode_id=periode.pk)
    
    try:
        try:
            wunsch_datum = datetime.strptime(datum_str, '%Y-%m-%d').date()
        except ValueError:
            wunsch_datum = datetime.strptime(datum_str, '%d. %B %Y').date()
    except ValueError:
        messages.error(request, f"❌ Ungültiges Datum: {datum_str}")
        return redirect('schichtplan:wunsch_kalender', periode_id=periode.pk)
    
    # Monat-Check
    if wunsch_datum.month != periode.fuer_monat.month or wunsch_datum.year != periode.fuer_monat.year:
        messages.error(request, "❌ Datum liegt nicht im Wunsch-Monat.")
        return redirect('schichtplan:wunsch_kalender', periode_id=periode.pk)
    
    # Lade oder erstelle Wunsch für den TARGET-Mitarbeiter
    wunsch, created = Schichtwunsch.objects.get_or_create(
        periode=periode,
        mitarbeiter=mitarbeiter,
        datum=wunsch_datum,
        defaults={'wunsch': 'tag_bevorzugt'}
    )
    
    # Speichern (POST)
    if request.method == 'POST':
        wunsch_kategorie = request.POST.get('wunsch')
        begruendung = request.POST.get('begruendung', '')
        mehrere_tage = request.POST.get('mehrere_tage') == 'on'
        
        # ========== NEU: VON-BIS LOGIK ==========
        if mehrere_tage and wunsch_kategorie in ['urlaub', 'krank']:
            von_datum_str = request.POST.get('von_datum')
            bis_datum_str = request.POST.get('bis_datum')
            
            # Validierung der Eingaben
            if not von_datum_str or not bis_datum_str:
                messages.error(request, "❌ Bitte beide Daten (Von und Bis) eingeben!")
                return redirect(request.path)
            
            try:
                von_datum = datetime.strptime(von_datum_str, '%Y-%m-%d').date()
                bis_datum = datetime.strptime(bis_datum_str, '%Y-%m-%d').date()
                
                # Validierung 1: Enddatum muss nach Startdatum liegen
                if bis_datum < von_datum:
                    messages.error(request, '❌ Das Enddatum muss nach dem Startdatum liegen!')
                    return redirect(request.path)
                
                # Validierung 2: Beide Daten müssen im Wunschmonat liegen
                if von_datum.month != periode.fuer_monat.month or von_datum.year != periode.fuer_monat.year:
                    messages.error(request, f"❌ Startdatum liegt nicht im Wunschmonat {periode.fuer_monat.strftime('%B %Y')}!")
                    return redirect(request.path)
                
                if bis_datum.month != periode.fuer_monat.month or bis_datum.year != periode.fuer_monat.year:
                    messages.error(request, f"❌ Enddatum liegt nicht im Wunschmonat {periode.fuer_monat.strftime('%B %Y')}!")
                    return redirect(request.path)
                
                # Validierung 3: Maximale Dauer prüfen (z.B. max. 31 Tage = ganzer Monat)
                differenz = (bis_datum - von_datum).days + 1
                if differenz > 31:
                    messages.error(request, '❌ Der Zeitraum darf maximal 31 Tage umfassen!')
                    return redirect(request.path)
                
                # Wünsche für alle Tage im Zeitraum erstellen
                erstellt = 0
                aktualisiert = 0
                fehler = 0
                
                aktuelles_datum = von_datum
                while aktuelles_datum <= bis_datum:
                    try:
                        wunsch_obj, created_flag = Schichtwunsch.objects.update_or_create(
                            periode=periode,
                            mitarbeiter=mitarbeiter,
                            datum=aktuelles_datum,
                            defaults={
                                'wunsch': wunsch_kategorie,
                                'begruendung': begruendung,
                                'benoetigt_genehmigung': True,  # NEU
                                'genehmigt': False
                            }
                        )
                        
                        if created_flag:
                            erstellt += 1
                        else:
                            aktualisiert += 1
                            
                    except Exception as e:
                        fehler += 1
                        print(f"Fehler bei Datum {aktuelles_datum}: {e}")
                    
                    aktuelles_datum += timedelta(days=1)
                
                # Erfolgsmeldung
                if fehler == 0:
                    messages.success(
                        request, 
                        f'✅ Urlaubswunsch für {mitarbeiter.schichtplan_kennung} erfolgreich eingetragen! '
                        f'Zeitraum: {von_datum.strftime("%d.%m.%Y")} - {bis_datum.strftime("%d.%m.%Y")} '
                        f'({differenz} Tage: {erstellt} neu, {aktualisiert} aktualisiert)'
                    )
                else:
                    messages.warning(
                        request,
                        f'⚠️ Urlaubswunsch teilweise eingetragen: {erstellt} neu, {aktualisiert} aktualisiert, {fehler} Fehler'
                    )
                
                return redirect('schichtplan:wunsch_kalender', periode_id=periode.pk)
                
            except ValueError as e:
                messages.error(request, f'❌ Ungültiges Datumsformat: {e}')
                return redirect(request.path)
        
        # ========== ORIGINAL: EINZELNER TAG ==========
        else:
            if wunsch_kategorie:
                wunsch.wunsch = wunsch_kategorie
                wunsch.begruendung = begruendung
                
                # Genehmigungsstatus setzen
                if wunsch_kategorie in ['urlaub', 'gar_nichts']:
                    wunsch.benoetigt_genehmigung = True
                    wunsch.genehmigt = False
                else:
                    wunsch.benoetigt_genehmigung = False
                    wunsch.genehmigt = True
                
                wunsch.save()
                
                messages.success(
                    request,
                    f"✅ Wunsch für {mitarbeiter.schichtplan_kennung} am {wunsch_datum.strftime('%d.%m.%Y')} gespeichert."
                )
                return redirect('schichtplan:wunsch_kalender', periode_id=periode.pk)
    
    # ========== GET-REQUEST: FORMULAR ANZEIGEN ==========
    
    # Andere Wünsche für diesen Tag laden (für Transparenz)
    andere_wuensche = Schichtwunsch.objects.filter(
        periode=periode,
        datum=wunsch_datum
    ).exclude(mitarbeiter=mitarbeiter).select_related('mitarbeiter')
    
    # Wunsch-Kategorien (aus deinem Model)
    wunsch_kategorien = Schichtwunsch.WUNSCH_KATEGORIEN
    
    context = {
        'periode': periode,
        'wunsch_datum': wunsch_datum,
        'wunsch': wunsch,
        'target_mitarbeiter': mitarbeiter,  # WICHTIG: für Template
        'andere_wuensche': andere_wuensche,
        'wunsch_kategorien': wunsch_kategorien,
        'is_planer': is_planer,
    }
    return render(request, 'schichtplan/wuensche_eingeben.html', context)
   
    
    # Andere Wünsche (Transparenz)
    andere_wuensche = Schichtwunsch.objects.filter(
        periode=periode,
        datum=wunsch_datum
    ).exclude(
        mitarbeiter=mitarbeiter # Schließe den MA aus, für den gerade eingetragen wird
    ).select_related('mitarbeiter').order_by('mitarbeiter__schichtplan_kennung')
    
    # Für Admins: Liste aller MA für ein evtl. Dropdown mitschicken
    alle_mitarbeiter = Mitarbeiter.objects.all().order_by('schichtplan_kennung') if is_planer else None

    context = {
        'periode': periode,
        'wunsch': wunsch,
        'wunsch_datum': wunsch_datum,
        'andere_wuensche': andere_wuensche,
        'wunsch_kategorien': Schichtwunsch.WUNSCH_KATEGORIEN,
        'is_planer': is_planer,
        'alle_mitarbeiter': alle_mitarbeiter,
        'target_mitarbeiter': mitarbeiter,
    }
    
    return render(request, 'schichtplan/wuensche_eingeben.html', context)


@login_required
def wunsch_ansehen(request, periode_id):
    """
    Zeigt ALLE Wünsche von MA1-MA15 für Transparenz.
    Mitarbeiter sehen Wünsche der Kollegen.
    Schichtplaner sehen alle + Genehmigungsstatus.
    """
    
    
    
    
    # Prüfe Berechtigung
    if not (ist_schichtplaner(request.user) or 
            (hasattr(request.user, 'mitarbeiter') and 
             request.user.mitarbeiter.schichtplan_kennung in 
             [f'MA{i}' for i in range(1, 16)])):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('arbeitszeit:dashboard')
    
    periode = get_object_or_404(SchichtwunschPeriode, pk=periode_id)
    
    # Generiere alle Tage
    jahr = periode.fuer_monat.year
    monat = periode.fuer_monat.month
    _, anzahl_tage = monthrange(jahr, monat)
    
    tage_im_monat = []
    for tag in range(1, anzahl_tage + 1):
        datum = datetime(jahr, monat, tag).date()
        tage_im_monat.append({
            'datum': datum,
            'wochentag': datum.strftime('%A'),
            'ist_wochenende': datum.weekday() >= 5,
        })
    
    # Lade alle Wünsche von MA1-MA15
    planbare_ma = get_planbare_mitarbeiter().order_by(
        Length('schichtplan_kennung'),
        'schichtplan_kennung'
    )
    
    alle_wuensche = Schichtwunsch.objects.filter(
        mitarbeiter__in=planbare_ma,
        datum__year=jahr,
        datum__month=monat
    ).select_related('mitarbeiter')
    
    # Gruppiere Wünsche: {datum: {mitarbeiter_id: wunsch_obj}}
    wuensche_matrix = defaultdict(dict)
    for wunsch in alle_wuensche:
        wuensche_matrix[wunsch.datum][wunsch.mitarbeiter.pk] = wunsch
    
    # Konflikt-Analyse (wie viele wollen Urlaub/frei pro Tag)
    konflikt_tage = {}
    for datum_obj in tage_im_monat:
        datum = datum_obj['datum']
        wuensche_am_tag = alle_wuensche.filter(datum=datum)
        
        urlaub_count = wuensche_am_tag.filter(wunsch='urlaub').count()
        frei_count = wuensche_am_tag.filter(wunsch='gar_nichts').count()
        
        gesamt_frei = urlaub_count + frei_count
        
        if gesamt_frei >= 3:  # Kritisch!
            konflikt_tage[datum] = {
                'urlaub': urlaub_count,
                'frei': frei_count,
                'gesamt': gesamt_frei,
                'level': 'danger' if gesamt_frei >= 5 else 'warning'
            }
    
    context = {
        'periode': periode,
        'tage_im_monat': tage_im_monat,
        'mitarbeiter_liste': planbare_ma,
        'wuensche_matrix': dict(wuensche_matrix),
        'konflikt_tage': konflikt_tage,
        'ist_planer': ist_schichtplaner(request.user),
        'eigener_mitarbeiter': request.user.mitarbeiter if hasattr(request.user, 'mitarbeiter') else None,
    }
    
    return render(request, 'schichtplan/wunsch_ansehen.html', context)

@login_required
def wuensche_genehmigen(request, periode_id):
    """
    Schichtplaner genehmigt Wünsche die Genehmigung benötigen.
    (Urlaub + gar_nichts)
    """
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('schichtplan:dashboard')
    
    periode = get_object_or_404(SchichtwunschPeriode, pk=periode_id)
    
    # Lade alle Wünsche die Genehmigung benötigen
    offene_wuensche = Schichtwunsch.objects.filter(
        periode=periode,
        benoetigt_genehmigung=True,
        genehmigt=False
    ).select_related('mitarbeiter').order_by('datum', 'mitarbeiter__schichtplan_kennung')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        wunsch_id = request.POST.get('wunsch_id')
        
        if action and wunsch_id:
            wunsch = get_object_or_404(Schichtwunsch, pk=wunsch_id)
            
            if action == 'genehmigen':
                wunsch.genehmigt = True
                wunsch.genehmigt_von = request.user
                wunsch.genehmigt_am = timezone.now()
                wunsch.save()
                
                messages.success(
                    request, 
                    f"✅ Wunsch von {wunsch.mitarbeiter.vollname} am {wunsch.datum} genehmigt!"
                )
            
            elif action == 'ablehnen':
                wunsch.delete()
                messages.info(
                    request,
                    f"ℹ️ Wunsch von {wunsch.mitarbeiter.vollname} am {wunsch.datum} abgelehnt."
                )
            
            return redirect('schichtplan:wuensche_genehmigen', periode_id=periode.pk)
    
    # Statistiken
    stats = {
        'offen': offene_wuensche.count(),
        'genehmigt': Schichtwunsch.objects.filter(
            periode=periode,
            genehmigt=True
        ).count(),
        'gesamt': Schichtwunsch.objects.filter(periode=periode).count(),
    }
    
    context = {
        'periode': periode,
        'offene_wuensche': offene_wuensche,
        'stats': stats,
    }
    
    return render(request, 'schichtplan/wuensche_genehmigen.html', context)

@login_required
def wunsch_kalender_aktuell(request):
    """Redirect auf die Kalenderansicht der aktuellen/ nächsten Wunschperiode oder aufs Dashboard."""
    from datetime import date
    heute = date.today()
    # Aktuelle oder nächste Periode (fuer_monat >= aktueller Monat)
    periode = (
        SchichtwunschPeriode.objects.filter(fuer_monat__gte=heute.replace(day=1))
        .order_by('fuer_monat')
        .first()
    )
    if periode:
        return redirect('schichtplan:wunsch_kalender', periode_id=periode.pk)
    # Keine passende Periode: zur Wunsch-Übersicht (Planer) oder Dashboard
    if ist_schichtplaner(request.user):
        return redirect('schichtplan:wunschperioden_liste')
    return redirect('arbeitszeit:dashboard')


@login_required
def wunsch_kalender(request, periode_id):
    """
    Kalenderansicht aller Wünsche einer Periode.
    TRANSPARENZ: Alle MA1-MA15 sehen die Wünsche der anderen!
    """
    
    import calendar
    
    
    # Berechtigungsprüfung
    if not hasattr(request.user, 'mitarbeiter'):
        messages.error(request, "❌ Kein Mitarbeiter-Profil gefunden.")
        return redirect('arbeitszeit:dashboard')
    
    mitarbeiter = request.user.mitarbeiter
    gueltige_kennungen = [f'MA{i}' for i in range(1, 16)]
    
    if not (ist_schichtplaner(request.user) or mitarbeiter.schichtplan_kennung in gueltige_kennungen):
        messages.error(request, "❌ Nur MA1-MA15 können Wünsche einsehen.")
        return redirect('arbeitszeit:dashboard')
    is_planer = ist_schichtplaner(request.user)
    
    # Lade Periode
    periode = get_object_or_404(SchichtwunschPeriode, pk=periode_id)
    
    # Lade ALLE MA1-MA15 und sortiere numerisch
    planbare_ma_queryset = get_planbare_mitarbeiter()
    
    # Sortiere Mitarbeiter numerisch nach MA-Nummer (MA1-MA15)
    def sortiere_ma_key(ma):
        kennung = ma.schichtplan_kennung or ''
        if kennung.startswith('MA') and len(kennung) > 2:
            try:
                nummer = int(kennung[2:])
                return nummer
            except ValueError:
                pass
        return 999  # Fallback für andere Kennungen
    
    planbare_ma = sorted(list(planbare_ma_queryset), key=sortiere_ma_key)
    
    # Lade alle Wünsche für diese Periode
    
    alle_wuensche = Schichtwunsch.objects.filter(
        periode=periode,
        mitarbeiter__in=planbare_ma_queryset
    ).select_related('mitarbeiter')
    
    # Gruppiere Wünsche nach Datum
    wuensche_nach_datum = defaultdict(list)
    for wunsch in alle_wuensche:
        wuensche_nach_datum[wunsch.datum].append(wunsch)
    
    # Erstelle Kalenderstruktur
    monat_start = date(periode.fuer_monat.year, periode.fuer_monat.month, 1)
    letzter_tag = calendar.monthrange(monat_start.year, monat_start.month)[1]
    monat_ende = date(monat_start.year, monat_start.month, letzter_tag)
    
    # Deutsche Wochentage und Monatsnamen
    WOCHENTAGE_DE = ('Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag', 'Samstag', 'Sonntag')
    MONATE_DE = ('', 'Januar', 'Februar', 'März', 'April', 'Mai', 'Juni', 'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember')
    
    # NRW-Feiertage im Monat
    feiertage_set, feiertage_namen = get_configured_feiertage(monat_start, monat_ende, region='nrw')
    
    kalender_daten = []
    current_date = monat_start
    
    while current_date <= monat_ende:
        tag_wuensche = wuensche_nach_datum.get(current_date, [])
        
        # Sortiere Wünsche nach MA-Nummer
        tag_wuensche_sortiert = sorted(tag_wuensche, key=lambda w: sortiere_ma_key(w.mitarbeiter))
        
        # Berechne Konflikte (Krank wird nicht mitgezählt)
        urlaube = sum(1 for w in tag_wuensche_sortiert if w.wunsch == 'urlaub')
        gar_nichts = sum(1 for w in tag_wuensche_sortiert if w.wunsch == 'gar_nichts')
        konflikt = None

        if urlaube + gar_nichts > 3:
            konflikt = {
                'typ': 'zu_viele_frei',
                'anzahl': urlaube + gar_nichts,
            }
        
        kalender_daten.append({
            'datum': current_date,
            'wochentag': WOCHENTAGE_DE[current_date.weekday()],
            'ist_wochenende': current_date.weekday() >= 5,
            'feiertag_name': feiertage_namen.get(current_date, ''),
            'wuensche': tag_wuensche_sortiert,
            'konflikt': konflikt,
            'hat_eigenen_wunsch': any(w.mitarbeiter == mitarbeiter for w in tag_wuensche_sortiert),
        })
        
        current_date += timedelta(days=1)
    
    # Beteiligung / Bewertung: Wie viele Wünsche hat jeder MA abgegeben?
    # Urlaub und Krank werden nicht mitgezählt
    wuensche_pro_ma = (
        alle_wuensche
        .exclude(wunsch__in=['urlaub', 'krank'])
        .values('mitarbeiter')
        .annotate(anzahl_tage=Count('datum', distinct=True))
        .order_by()
    )
    ma_anzahl = {e['mitarbeiter']: e['anzahl_tage'] for e in wuensche_pro_ma}
    beteiligung_liste = []
    for ma in planbare_ma:
        n = ma_anzahl.get(ma.pk, 0)
        if n == 0:
            einordnung = 'keine'
            einordnung_label = 'Keine Wünsche abgegeben'
        elif n <= 4:
            einordnung = 'wenige'
            einordnung_label = 'Wenige Angaben'
        elif n <= 14:
            einordnung = 'einige'
            einordnung_label = 'Mittlere Beteiligung'
        else:
            einordnung = 'viele'
            einordnung_label = 'Viele Wünsche'
        beteiligung_liste.append({
            'ma': ma,
            'anzahl_tage': n,
            'einordnung': einordnung,
            'einordnung_label': einordnung_label,
        })
    
    context = {
        'periode': periode,
        'is_planer': is_planer,
        'kalender_daten': kalender_daten,
        'mitarbeiter': mitarbeiter,
        'alle_mitarbeiter': planbare_ma,
        'monat_name': MONATE_DE[monat_start.month],
        'beteiligung_liste': beteiligung_liste,
    }
    
    return render(request, 'schichtplan/wunsch_kalender.html', context)


@login_required
def wunsch_schnell_setzen(request, periode_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Nur POST erlaubt.'}, status=405)

    if not hasattr(request.user, 'mitarbeiter'):
        return JsonResponse({'error': 'Kein Mitarbeiter-Profil gefunden.'}, status=403)

    periode = get_object_or_404(SchichtwunschPeriode, pk=periode_id)
    if not periode.ist_offen:
        return JsonResponse({'error': 'Wunschperiode ist geschlossen.'}, status=403)

    mitarbeiter = request.user.mitarbeiter
    ma_id = request.POST.get('mitarbeiter_id')
    if not ma_id or str(mitarbeiter.pk) != str(ma_id):
        return JsonResponse({'error': 'Nur eigene Spalte erlaubt.'}, status=403)

    wunsch = request.POST.get('wunsch')
    erlaubte = {
        'urlaub',
        'kein_tag_aber_nacht',
        'keine_nacht_aber_tag',
        'tag_bevorzugt',
        'nacht_bevorzugt',
        'gar_nichts',
        'zusatzarbeit',
    }
    if wunsch not in erlaubte:
        return JsonResponse({'error': 'Unbekannter Wunsch.'}, status=400)

    def set_wunsch_for_date(datum):
        if datum.month != periode.fuer_monat.month or datum.year != periode.fuer_monat.year:
            raise ValueError('Datum liegt nicht im Wunschmonat.')
        benoetigt = wunsch in ['urlaub', 'gar_nichts']
        Schichtwunsch.objects.update_or_create(
            periode=periode,
            mitarbeiter=mitarbeiter,
            datum=datum,
            defaults={
                'wunsch': wunsch,
                'begruendung': '',
                'benoetigt_genehmigung': benoetigt,
                'genehmigt': False if benoetigt else True,
            }
        )

    von_datum_str = request.POST.get('von_datum')
    bis_datum_str = request.POST.get('bis_datum')
    if wunsch in ['urlaub', 'krank'] and von_datum_str and bis_datum_str:
        try:
            von_datum = datetime.strptime(von_datum_str, '%Y-%m-%d').date()
            bis_datum = datetime.strptime(bis_datum_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Ungueltiges Datumsformat.'}, status=400)

        if bis_datum < von_datum:
            return JsonResponse({'error': 'Enddatum muss nach Startdatum liegen.'}, status=400)
        if von_datum.month != periode.fuer_monat.month or von_datum.year != periode.fuer_monat.year:
            return JsonResponse({'error': 'Startdatum liegt nicht im Wunschmonat.'}, status=400)
        if bis_datum.month != periode.fuer_monat.month or bis_datum.year != periode.fuer_monat.year:
            return JsonResponse({'error': 'Enddatum liegt nicht im Wunschmonat.'}, status=400)

        count = 0
        current = von_datum
        while current <= bis_datum:
            set_wunsch_for_date(current)
            count += 1
            current += timedelta(days=1)
        return JsonResponse({'ok': True, 'count': count})

    datum_str = request.POST.get('datum')
    if not datum_str:
        return JsonResponse({'error': 'Kein Datum angegeben.'}, status=400)
    try:
        datum = datetime.strptime(datum_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Ungueltiges Datumsformat.'}, status=400)

    try:
        set_wunsch_for_date(datum)
    except ValueError as exc:
        return JsonResponse({'error': str(exc)}, status=400)

    return JsonResponse({'ok': True})
    
    


@login_required
def wunsch_loeschen(request, wunsch_id):
    # 1. Wunsch holen
    wunsch = get_object_or_404(Schichtwunsch, pk=wunsch_id)
    periode_id = wunsch.periode.pk if wunsch.periode else None
    next_uebersicht = request.GET.get('next') == 'uebersicht_detail'
    schichtplan_pk = request.GET.get('schichtplan_pk', '').strip()

    # 2. Berechtigung: Schichtplaner-Funktion aufrufen
    ist_planer = ist_schichtplaner(request.user)

    # 3. Berechtigung: Besitzer
    ist_besitzer = False
    if hasattr(request.user, 'mitarbeiter') and wunsch.mitarbeiter:
        if wunsch.mitarbeiter == request.user.mitarbeiter:
            ist_besitzer = True

    # 4. SICHERHEITS-ABBRUCH
    darf_krank_loeschen = wunsch.wunsch == 'krank'
    if not (ist_planer or ist_besitzer or darf_krank_loeschen):
        messages.error(request, f"❌ Keine Berechtigung. (Planer-Status: {ist_planer})")
        if next_uebersicht and schichtplan_pk:
            return redirect('schichtplan:uebersicht_detail', pk=schichtplan_pk)
        if periode_id:
            return redirect('schichtplan:wunsch_kalender', periode_id=periode_id)
        return redirect('schichtplan:dashboard')

    # 5. DER LÖSCHVORGANG
    if request.method == 'POST':
        ma_kennung = wunsch.mitarbeiter.schichtplan_kennung if wunsch.mitarbeiter else "Unbekannt"
        wunsch.delete()
        if ist_planer and not ist_besitzer:
            messages.success(request, f"✅ Schichtplaner-Aktion: Wunsch für {ma_kennung} wurde gelöscht.")
        else:
            messages.success(request, "✅ Ihr Wunsch wurde gelöscht.")
        if next_uebersicht and schichtplan_pk:
            return redirect('schichtplan:uebersicht_detail', pk=schichtplan_pk)
        if periode_id:
            return redirect('schichtplan:wunsch_kalender', periode_id=periode_id)
        return redirect('schichtplan:dashboard')

    # Falls GET: Bestätigungsseite anzeigen
    return render(request, 'schichtplan/wunsch_loeschen_confirm.html', {
        'wunsch': wunsch,
        'ist_planer': ist_planer,
        'next_uebersicht': next_uebersicht,
        'schichtplan_pk': schichtplan_pk,
    })

@login_required
def wuensche_schichtplaner_uebersicht(request, periode_id):
    """
    Übersicht aller Wünsche für Schichtplaner.
    Zeigt Statistiken und Liste aller Wünsche von MA1-MA15.
    """
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung für diese Seite.")
        return redirect('arbeitszeit:dashboard')
    
    
 
    
    periode = get_object_or_404(SchichtwunschPeriode, pk=periode_id)
    
    # Lade alle Wünsche von MA1-MA15
    planbare_ma = get_planbare_mitarbeiter()
    
    alle_wuensche = Schichtwunsch.objects.filter(
        periode=periode,
        mitarbeiter__in=planbare_ma
    ).select_related('mitarbeiter').order_by('datum', 'mitarbeiter__schichtplan_kennung')
    
    # Statistiken
    stats = {
        'gesamt': alle_wuensche.count(),
        'urlaube': alle_wuensche.filter(wunsch='urlaub').count(),
        'gar_nichts': alle_wuensche.filter(wunsch='gar_nichts').count(),
        'offen_genehmigung': alle_wuensche.filter(
            benoetigt_genehmigung=True,
            genehmigt=False
        ).count(),
        'genehmigt': alle_wuensche.filter(genehmigt=True).count(),
    }
    
    # Gruppiere nach Mitarbeiter
    wuensche_nach_ma = defaultdict(list)
    for wunsch in alle_wuensche:
        wuensche_nach_ma[wunsch.mitarbeiter].append(wunsch)
    
    # Sortiere Mitarbeiter numerisch nach MA-Nummer (MA1-MA15)
    def sortiere_ma(ma):
        kennung = ma.schichtplan_kennung or ''
        if kennung.startswith('MA') and len(kennung) > 2:
            try:
                nummer = int(kennung[2:])
                return nummer
            except ValueError:
                pass
        return 999  # Fallback für andere Kennungen
    
    sortierte_ma = sorted(wuensche_nach_ma.keys(), key=sortiere_ma)
    
    context = {
        'periode': periode,
        'wuensche_nach_ma': dict(wuensche_nach_ma),
        'sortierte_ma': sortierte_ma,
        'stats': stats,
        'alle_mitarbeiter': planbare_ma,
    }
    
    return render(request, 'schichtplan/wuensche_schichtplaner_uebersicht.html', context)


@login_required
def wunsch_genehmigen(request, wunsch_id):
    """
    Genehmigt oder lehnt einen einzelnen Wunsch ab.
    Nur für Schichtplaner.
    """
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('arbeitszeit:dashboard')
    
    
    
    wunsch = get_object_or_404(Schichtwunsch, pk=wunsch_id)
    
    if request.method == 'POST':
        aktion = request.POST.get('aktion')
        
        if aktion == 'genehmigen':
            wunsch.genehmigt = True
            wunsch.genehmigt_von = request.user
            wunsch.genehmigt_am = timezone.now()
            wunsch.save()
            
            messages.success(
                request,
                f"✅ Wunsch von {wunsch.mitarbeiter.vollname} wurde genehmigt."
            )
        
        elif aktion == 'ablehnen':
            # Ablehnen = Wunsch löschen oder Flag setzen
            wunsch.genehmigt = False
            wunsch.genehmigt_von = None
            wunsch.genehmigt_am = None
            wunsch.save()
            
            messages.info(
                request,
                f"ℹ️ Wunsch von {wunsch.mitarbeiter.vollname} wurde abgelehnt."
            )
        
        # Redirect zurück zur Übersicht
        if wunsch.periode:
            return redirect('schichtplan:wuensche_schichtplaner_uebersicht', periode_id=wunsch.periode.pk)
        else:
            return redirect('schichtplan:dashboard')
    
    # GET: Zeige Genehmigungsformular
    context = {
        'wunsch': wunsch,
    }
    
    return render(request, 'schichtplan/wunsch_genehmigen.html', context)

@login_required
def wunschperioden_liste(request):
    # RIEGEL VOR: Nur Planer dürfen die Liste aller Perioden sehen
    if not ist_schichtplaner(request.user):
        messages.error(request, "Kein Zugriff auf die Perioden-Verwaltung.")
        return redirect('schichtplan:wunsch_kalender_aktuell')

    perioden = SchichtwunschPeriode.objects.all().order_by('-fuer_monat')
    return render(request, 'schichtplan/wunsch_perioden_liste.html', {
        'perioden': perioden,
        'is_planer': True
    })

@login_required
def wunschperiode_loeschen(request, periode_id):
    """
    Löscht eine Wunschperiode inkl. aller zugehörigen Wünsche.
    Nur für Schichtplaner.
    """
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('arbeitszeit:dashboard')
    
    periode = get_object_or_404(SchichtwunschPeriode, pk=periode_id)
    
    if request.method == 'POST':
        # Zähle Wünsche vor dem Löschen
        anzahl_wuensche = Schichtwunsch.objects.filter(periode=periode).count()
        periode_name = periode.name
        
        # Lösche Periode (CASCADE löscht automatisch alle Wünsche)
        periode.delete()
        
        messages.success(
            request,
            f"✅ Wunschperiode '{periode_name}' wurde gelöscht ({anzahl_wuensche} Wünsche entfernt)."
        )
        return redirect('schichtplan:wunschperioden_liste')
    
    # GET: Zeige Bestätigungsseite
    anzahl_wuensche = Schichtwunsch.objects.filter(periode=periode).count()
    
    context = {
        'periode': periode,
        'anzahl_wuensche': anzahl_wuensche,
    }
    
    return render(request, 'schichtplan/wunschperiode_loeschen_confirm.html', context)

# Hilfsfunktion: Ist User ein Schichtplaner?
def ist_schichtplaner(user):
    """Prüft, ob User Schichtplaner-Rechte hat"""
    # ANPASSEN: Je nach Ihrer Berechtigungslogik
    return user.is_staff or user.groups.filter(name='Schichtplaner').exists()


@login_required
def genehmigungen_uebersicht(request):
    """
    Übersicht aller Perioden mit offenen Genehmigungen.
    Nur für Schichtplaner zugänglich.
    """
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung für Urlaubsgenehmigungen.")
        return redirect('arbeitszeit:dashboard')
    
    # Alle Perioden mit offenen Urlaubsanträgen
    perioden = SchichtwunschPeriode.objects.annotate(
        offene_urlaube=Count(
            'schichtwunsch',
            filter=Q(
                schichtwunsch__wunsch__in=['urlaub', 'gar_nichts'],
                schichtwunsch__genehmigt=False
            )
        )
    ).filter(offene_urlaube__gt=0).order_by('-fuer_monat')
    
    context = {
        'perioden': perioden,
    }
    
    return render(request, 'schichtplan/genehmigungen_uebersicht.html', context)


@login_required
def genehmigungen_periode(request, periode_id):
    """
    Detailansicht: Alle Urlaubsanträge einer Periode mit Konflikt-Analyse.
    """
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('arbeitszeit:dashboard')
    
    periode = get_object_or_404(SchichtwunschPeriode, pk=periode_id)
    
    # Alle Urlaubsanträge (offen + genehmigt)
    alle_urlaube = Schichtwunsch.objects.filter(
        periode=periode,
        wunsch__in=['urlaub', 'gar_nichts']
    ).select_related('mitarbeiter').order_by('datum', 'mitarbeiter__schichtplan_kennung')
    
    # Offene Anträge
    offene_antraege = alle_urlaube.filter(genehmigt=False)
    
    # Genehmigte Anträge
    genehmigte_antraege = alle_urlaube.filter(genehmigt=True)
    
    # ====================================================================
    # KONFLIKT-ANALYSE: Welche Tage sind kritisch?
    # ====================================================================
    
    # Zähle genehmigte Urlaube pro Tag
    urlaube_pro_tag = {}
    for urlaub in genehmigte_antraege:
        datum = urlaub.datum
        if datum not in urlaube_pro_tag:
            urlaube_pro_tag[datum] = 0
        urlaube_pro_tag[datum] += 1
    
    # Gruppiere offene Anträge nach Datum
    antraege_nach_datum = {}
    for antrag in offene_antraege:
        datum = antrag.datum
        if datum not in antraege_nach_datum:
            antraege_nach_datum[datum] = {
                'datum': datum,
                'antraege': [],
                'bereits_genehmigt': urlaube_pro_tag.get(datum, 0),
                'verfuegbar_jetzt': 15 - urlaube_pro_tag.get(datum, 0),
                'status': 'ok'  # ok, warnung, kritisch
            }
        antraege_nach_datum[datum]['antraege'].append(antrag)
    
    # Status berechnen
    for datum_str, daten in antraege_nach_datum.items():
        bereits = daten['bereits_genehmigt']
        offen = len(daten['antraege'])
        gesamt_wenn_alle = bereits + offen
        verfuegbar_dann = 15 - gesamt_wenn_alle
        
        daten['verfuegbar_wenn_alle'] = verfuegbar_dann
        
        # Status setzen
        if verfuegbar_dann < 4:
            daten['status'] = 'kritisch'
        elif verfuegbar_dann < 6:
            daten['status'] = 'warnung'
        else:
            daten['status'] = 'ok'
    
    # Sortiere nach Datum
    antraege_nach_datum = dict(sorted(antraege_nach_datum.items()))
    
    # Statistiken
    kritische_tage = sum(1 for d in antraege_nach_datum.values() if d['status'] == 'kritisch')
    warn_tage = sum(1 for d in antraege_nach_datum.values() if d['status'] == 'warnung')
    
    context = {
        'periode': periode,
        'antraege_nach_datum': antraege_nach_datum,
        'offene_antraege_count': offene_antraege.count(),
        'genehmigte_antraege_count': genehmigte_antraege.count(),
        'kritische_tage': kritische_tage,
        'warn_tage': warn_tage,
    }
    
    return render(request, 'schichtplan/genehmigungen_periode.html', context)


@login_required
def urlaub_genehmigen(request, wunsch_id):
    """
    Genehmigt einen einzelnen Urlaubsantrag mit Konfliktprüfung.
    """
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('arbeitszeit:dashboard')
    
    wunsch = get_object_or_404(Schichtwunsch, pk=wunsch_id)
    
    if request.method == 'POST':
        # Prüfe, wie viele Urlaube an diesem Tag bereits genehmigt sind
        urlaube_am_tag = Schichtwunsch.objects.filter(
            periode=wunsch.periode,
            datum=wunsch.datum,
            wunsch__in=['urlaub', 'gar_nichts'],
            genehmigt=True
        ).count()
        
        verfuegbar_nach_genehmigung = 15 - (urlaube_am_tag + 1)
        
        # KRITISCH: Wenn < 4 MA verfügbar wären
        if verfuegbar_nach_genehmigung < 4:
            messages.error(
                request,
                f"❌ Genehmigung nicht möglich! Am {wunsch.datum.strftime('%d.%m.%Y')} würden nur noch "
                f"{verfuegbar_nach_genehmigung} MA verfügbar sein (benötigt: mindestens 4)."
            )
            return redirect('schichtplan:genehmigungen_periode', periode_id=wunsch.periode.pk)
        
        # WARNUNG: Wenn < 6 MA verfügbar wären
        elif verfuegbar_nach_genehmigung < 6:
            messages.warning(
                request,
                f"⚠️ Achtung: Am {wunsch.datum.strftime('%d.%m.%Y')} werden nur noch "
                f"{verfuegbar_nach_genehmigung} MA verfügbar sein (sehr eng!)."
            )
        
        # Genehmigung durchführen
        wunsch.genehmigt = True
        wunsch.genehmigt_von = request.user
        wunsch.genehmigt_am = timezone.now()
        wunsch.benoetigt_genehmigung = True
        wunsch.save()
        
        messages.success(
            request,
            f"✅ Urlaub für {wunsch.mitarbeiter.schichtplan_kennung} am "
            f"{wunsch.datum.strftime('%d.%m.%Y')} genehmigt."
        )
        
        return redirect('schichtplan:genehmigungen_periode', periode_id=wunsch.periode.pk)
    
    # GET: Zeige Bestätigungsseite
    # Berechne Verfügbarkeit
    urlaube_am_tag = Schichtwunsch.objects.filter(
        periode=wunsch.periode,
        datum=wunsch.datum,
        wunsch__in=['urlaub', 'gar_nichts'],
        genehmigt=True
    ).count()
    
    verfuegbar_nach_genehmigung = 15 - (urlaube_am_tag + 1)
    
    context = {
        'wunsch': wunsch,
        'urlaube_am_tag': urlaube_am_tag,
        'verfuegbar_nach_genehmigung': verfuegbar_nach_genehmigung,
    }
    
    return render(request, 'schichtplan/urlaub_genehmigen_confirm.html', context)


@login_required
def urlaub_ablehnen(request, wunsch_id):
    """
    Lehnt einen Urlaubsantrag ab.
    """
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('arbeitszeit:dashboard')
    
    wunsch = get_object_or_404(Schichtwunsch, pk=wunsch_id)
    
    if request.method == 'POST':
        # Optional: Ablehnungsgrund speichern
        ablehnungsgrund = request.POST.get('grund', '')
        
        # Wunsch löschen oder als abgelehnt markieren
        # Option A: Löschen
        mitarbeiter_name = wunsch.mitarbeiter.schichtplan_kennung
        datum_str = wunsch.datum.strftime('%d.%m.%Y')
        periode_id = wunsch.periode.pk
        
        wunsch.delete()
        
        messages.info(
            request,
            f"❌ Urlaubsantrag von {mitarbeiter_name} für {datum_str} abgelehnt."
        )
        
        # Option B: Als abgelehnt markieren (wenn Sie ein 'abgelehnt' Feld haben)
        # wunsch.abgelehnt = True
        # wunsch.ablehnungsgrund = ablehnungsgrund
        # wunsch.save()
        
        return redirect('schichtplan:genehmigungen_periode', periode_id=periode_id)
    
    # GET: Bestätigungsseite
    context = {
        'wunsch': wunsch,
    }
    
    return render(request, 'schichtplan/urlaub_ablehnen_confirm.html', context)

@login_required
def urlaub_bulk_genehmigen(request):
    """
    Genehmigt mehrere Urlaubsanträge auf einmal.
    """
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('arbeitszeit:dashboard')
    
    if request.method != 'POST':
        messages.error(request, "❌ Ungültige Anfrage.")
        return redirect('schichtplan:genehmigungen_uebersicht')
    
    # IDs der ausgewählten Anträge aus POST-Daten
    wunsch_ids = request.POST.getlist('wunsch_ids[]')
    
    if not wunsch_ids:
        messages.warning(request, "⚠️ Keine Anträge ausgewählt.")
        return redirect(request.META.get('HTTP_REFERER', 'schichtplan:genehmigungen_uebersicht'))
    
    # Anträge laden
    wuensche = Schichtwunsch.objects.filter(
        id__in=wunsch_ids,
        genehmigt=False
    ).select_related('mitarbeiter', 'periode')
    
    if not wuensche:
        messages.warning(request, "⚠️ Keine gültigen Anträge gefunden.")
        return redirect(request.META.get('HTTP_REFERER', 'schichtplan:genehmigungen_uebersicht'))
    
    # Gruppiere nach Datum für Konflikt-Check
    wuensche_nach_datum = {}
    for w in wuensche:
        if w.datum not in wuensche_nach_datum:
            wuensche_nach_datum[w.datum] = []
        wuensche_nach_datum[w.datum].append(w)
    
    # Konflikt-Prüfung für jeden Tag
    kritische_tage = []
    for datum, tag_wuensche in wuensche_nach_datum.items():
        periode = tag_wuensche[0].periode
        
        # Bereits genehmigte Urlaube an diesem Tag
        bereits_genehmigt = Schichtwunsch.objects.filter(
            periode=periode,
            datum=datum,
            wunsch__in=['urlaub', 'gar_nichts'],
            genehmigt=True
        ).count()
        
        # Wie viele würden genehmigt werden?
        neue_genehmigungen = len(tag_wuensche)
        
        # Verfügbare MA nach Genehmigung
        verfuegbar_nach = 15 - (bereits_genehmigt + neue_genehmigungen)
        
        if verfuegbar_nach < 4:
            kritische_tage.append({
                'datum': datum,
                'verfuegbar': verfuegbar_nach,
                'anzahl_antraege': neue_genehmigungen
            })
    
    # Wenn kritische Tage: ABBRUCH
    if kritische_tage:
        tage_str = ", ".join([k['datum'].strftime('%d.%m.%Y') for k in kritische_tage])
        messages.error(
            request,
            f"❌ Bulk-Genehmigung abgebrochen!\n\n"
            f"An folgenden Tagen würde die Mindestbesetzung unterschritten:\n"
            f"{tage_str}\n\n"
            f"Bitte genehmigen Sie diese Tage einzeln oder wählen Sie weniger Anträge aus."
        )
        return redirect(request.META.get('HTTP_REFERER', 'schichtplan:genehmigungen_uebersicht'))
    
    # Alle genehmigen
    genehmigt = 0
    for wunsch in wuensche:
        wunsch.genehmigt = True
        wunsch.genehmigt_von = request.user
        wunsch.genehmigt_am = timezone.now()
        wunsch.benoetigt_genehmigung = True
        wunsch.save()
        genehmigt += 1
    
    messages.success(
        request,
        f"✅ {genehmigt} Urlaubsanträge erfolgreich genehmigt!"
    )
    
    return redirect(request.META.get('HTTP_REFERER', 'schichtplan:genehmigungen_uebersicht'))


@login_required
def urlaub_bulk_ablehnen(request):
    """
    Lehnt mehrere Urlaubsanträge auf einmal ab.
    """
    if not ist_schichtplaner(request.user):
        messages.error(request, "❌ Keine Berechtigung.")
        return redirect('arbeitszeit:dashboard')
    
    if request.method != 'POST':
        messages.error(request, "❌ Ungültige Anfrage.")
        return redirect('schichtplan:genehmigungen_uebersicht')
    
    # IDs der ausgewählten Anträge
    wunsch_ids = request.POST.getlist('wunsch_ids[]')
    
    if not wunsch_ids:
        messages.warning(request, "⚠️ Keine Anträge ausgewählt.")
        return redirect(request.META.get('HTTP_REFERER', 'schichtplan:genehmigungen_uebersicht'))
    
    # Anträge löschen
    geloescht = Schichtwunsch.objects.filter(
        id__in=wunsch_ids,
        genehmigt=False
    ).delete()[0]
    
    if geloescht > 0:
        messages.success(
            request,
            f"✅ {geloescht} Urlaubsanträge abgelehnt."
        )
    else:
        messages.warning(request, "⚠️ Keine gültigen Anträge gefunden.")
    
    return redirect(request.META.get('HTTP_REFERER', 'schichtplan:genehmigungen_uebersicht'))
