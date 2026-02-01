import openpyxl
from django.utils.dateparse import parse_date
from datetime import datetime
from .models import Schicht, Schichtplan, Schichttyp, Mitarbeiter

def import_schichtplan(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    
    # Beispiel: Ein Schichtplan für das Jahr (z. B. Januar 2026)
    schichtplan_name = f"Schichtplan {datetime.now().strftime('%B %Y')}"
    schichtplan = Schichtplan.objects.create(
        name=schichtplan_name,
        start_datum="2025-11-01",  # Beispielstartdatum
        ende_datum="2025-11-30",    # Beispielenddatum
    )
    
    # Schichttypen holen (z. B. Tag, Nacht)
    tag_schicht = Schichttyp.objects.get(kuerzel='T')
    nacht_schicht = Schichttyp.objects.get(kuerzel='N')
    
    # Durch die Zeilen der Excel-Datei gehen und die Schichten importieren
    for row in sheet.iter_rows(min_row=2, max_row=30, min_col=2, max_col=16, values_only=True):
        datum = parse_date(f"2026-01-{sheet.cell(row=row[0], column=1).row}")  # Datum durch Zeile holen
        
        for idx, value in enumerate(row):
            if value == 'T':  # Tagschicht
                Schicht.objects.create(
                    schichtplan=schichtplan,
                    mitarbeiter=Mitarbeiter.objects.get(id=idx+1),
                    datum=datum,
                    schichttyp=tag_schicht
                )
            elif value == 'N':  # Nachtschicht
                Schicht.objects.create(
                    schichtplan=schichtplan,
                    mitarbeiter=Mitarbeiter.objects.get(id=idx+1),
                    datum=datum,
                    schichttyp=nacht_schicht
                )
            elif value == 'U':  # Urlaub
                # Urlaubsdaten oder eine andere Verarbeitung für den Urlaub
                pass
            # Hier kannst du auch Zusatzarbeit oder andere Arten von Schichten verarbeiten
            # elif value == 'Z':  # Zusatzarbeit
            #    ...
    
    return schichtplan